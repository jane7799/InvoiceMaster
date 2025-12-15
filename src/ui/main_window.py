
import os
import sys
import gc
import logging
import platform
import fitz  # PyMuPDF
import pandas as pd
from PyQt6.QtWidgets import (QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                           QLabel, QPushButton, QListWidget, QListWidgetItem, 
                           QStackedWidget, QComboBox, QCheckBox, QRadioButton,
                           QButtonGroup, QToolButton, QFileDialog, QMessageBox,
                           QInputDialog, QSpinBox, QApplication, QAbstractItemView,
                           QMenu)
from PyQt6.QtCore import Qt, QSize, QTimer
from PyQt6.QtGui import QIcon, QImage, QAction, QTransform
from PyQt6.QtPrintSupport import QPrinter, QPrinterInfo, QPrintDialog
from PyQt6.QtCore import QSettings

from src.core.invoice_helper import InvoiceHelper
from src.core.pdf_engine import PDFEngine
from src.core.workers import OcrWorker, PdfWorker, PrintWorker
from src.core.license_manager import LicenseManager
from src.core.database import get_db
from src.themes.theme_manager import ThemeManager
from src.utils.log_manager import LogManager
from src.utils.icons import Icons
from src.utils.constants import APP_NAME, APP_VERSION, APP_AUTHOR_CN
from src.utils.utils import resource_path
from src.utils.config import UI_CONFIG

from src.ui.dialogs import ProgressDialog, AboutDialog, ActivationDialog
from src.ui.settings_dialog import SettingsDlg
from src.ui.statistics_dialog import StatisticsDialog
from src.ui.widgets import Card, DragArea, InvoiceItemWidget
from src.ui.preview import AdvancedPreviewArea, SingleDocViewer

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} {APP_VERSION}")
        self.resize(1350, 850); self.data = []; self.theme_c = "#555"
        self.temp_files = [] 
        self.preview_timer = QTimer(); self.preview_timer.setSingleShot(True); self.preview_timer.timeout.connect(self.generate_realtime_preview)
        self.current_printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        self.right_panel = None; self.settings_card = None
        self.license_manager = LicenseManager()
        
        # å¼‚æ­¥å·¥ä½œçº¿ç¨‹å¼•ç”¨
        self.ocr_worker = None
        self.pdf_worker = None
        self.print_worker = None
        self.progress_dialog = None
        
        self.init_ui()
        ThemeManager.apply(QApplication.instance())
        self.change_theme("Light")

    def closeEvent(self, event):
        """åº”ç”¨å…³é—­æ—¶æ¸…ç†èµ„æº"""
        import logging
        logger = logging.getLogger(__name__)
        
        # æ¸…ç†ä¸´æ—¶æ–‡ä»¶
        for f in self.temp_files:
            try: os.remove(f)
            except: pass
        
        # [V3.6] æ¸…é™¤æ•°æ®åº“ç¼“å­˜ï¼Œé¿å…ä¸Šä¸€æ¬¡æ•°æ®å¸¦å…¥ä¸‹ä¸€æ¬¡
        try:
            from src.core.database import get_db
            deleted = get_db().clear_all()
            logger.info(f"å…³é—­æ—¶æ¸…é™¤æ•°æ®åº“ç¼“å­˜: {deleted} æ¡è®°å½•")
        except Exception as e:
            logger.warning(f"æ¸…é™¤æ•°æ®åº“ç¼“å­˜å¤±è´¥: {e}")
        
        super().closeEvent(event)

    def init_ui(self):
        main = QWidget(); self.setCentralWidget(main)
        layout = QHBoxLayout(main); layout.setContentsMargins(15,15,15,15); layout.setSpacing(15)

        # LEFT
        left = QWidget(); left.setFixedWidth(280); lv = QVBoxLayout(left); lv.setContentsMargins(0,0,0,0); lv.setSpacing(12)
        
        # æ‹–æ”¾åŒºåŸŸ
        self.drag = DragArea(); self.drag.dropped.connect(self.add_files)
        lv.addWidget(self.drag)
        
        # å‘ç¥¨æ¸…å•æ ‡é¢˜
        list_title = QLabel("ğŸ“‹ å‘ç¥¨æ¸…å• (åŒå‡»ä¿®æ­£é‡‘é¢)")
        list_title.setStyleSheet("color: #64748B; font-size: 12px; font-weight: 600; margin-top: 8px;")
        lv.addWidget(list_title)
        
        self.list = QListWidget(); self.list.setIconSize(QSize(40,50)); self.list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu); self.list.customContextMenuRequested.connect(self.ctx_menu)
        self.list.itemDoubleClicked.connect(self.edit_item); self.list.itemClicked.connect(self.show_single_doc)
        
        tb = QHBoxLayout(); tb.setSpacing(10)
        self.btn_set = QPushButton("è®¾ç½®")
        self.btn_set.setMinimumHeight(44)
        # å¹³å°è‡ªé€‚åº”æŒ‰é’®æ ·å¼
        if UI_CONFIG.get("use_gradients", True):
            btn_bg = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3B82F6, stop:1 #2563EB);"
            btn_hover = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2563EB, stop:1 #1D4ED8);"
        else:
            btn_bg = "background: #2563EB;"
            btn_hover = "background: #1D4ED8;"
        self.btn_set.setStyleSheet(f"""
            QPushButton {{
                {btn_bg}
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                {btn_hover}
            }}
        """)
        self.btn_set.clicked.connect(lambda: SettingsDlg(self).exec())
        
        self.btn_del = QPushButton("æ¸…ç©º")
        self.btn_del.setMinimumHeight(44)
        if UI_CONFIG.get("use_gradients", True):
            del_bg = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #60A5FA, stop:1 #3B82F6);"
            del_hover = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3B82F6, stop:1 #2563EB);"
        else:
            del_bg = "background: #3B82F6;"
            del_hover = "background: #2563EB;"
        self.btn_del.setStyleSheet(f"""
            QPushButton {{
                {del_bg}
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                {del_hover}
            }}
        """)
        self.btn_del.clicked.connect(self.clear)
        
        # ç»Ÿè®¡æŒ‰é’®
        self.btn_stats = QPushButton("ğŸ“Š ç»Ÿè®¡")
        self.btn_stats.setMinimumHeight(44)
        self.btn_stats.setStyleSheet(f"""
            QPushButton {{
                background: #10B981;
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                background: #059669;
            }}
        """)
        self.btn_stats.clicked.connect(lambda: StatisticsDialog(self, self.data).exec())
        
        tb.addWidget(self.btn_set); tb.addWidget(self.btn_stats); tb.addStretch(); tb.addWidget(self.btn_del)
        
        lv.addWidget(self.list); lv.addLayout(tb)
        footer_lbl = QLabel(APP_AUTHOR_CN, alignment=Qt.AlignmentFlag.AlignCenter); footer_lbl.setStyleSheet("color:#999; font-size:11px; margin-top: 10px;")
        lv.addWidget(footer_lbl)

        # MIDDLE
        mid = QWidget(); mv = QVBoxLayout(mid); mv.setContentsMargins(0,0,0,0); mv.setSpacing(0)
        self.stack = QStackedWidget()
        self.word_preview = AdvancedPreviewArea() 
        self.single_viewer = SingleDocViewer() 
        self.stack.addWidget(self.word_preview); self.stack.addWidget(self.single_viewer)
        mv.addWidget(self.stack)

        # RIGHT
        self.right_panel = QWidget(); self.right_panel.setFixedWidth(340)
        rv = QVBoxLayout(self.right_panel); rv.setContentsMargins(0,0,0,0)
        self.settings_card = Card() 
        self.settings_layout = QVBoxLayout(self.settings_card)
        self.settings_layout.setSpacing(15); self.settings_layout.setContentsMargins(20,20,20,20)
        
        # æ‰“å°è®¾ç½®æ ‡é¢˜
        print_title = QLabel("ğŸ–¨ï¸ æ‰“å°è®¾ç½®")
        print_title.setStyleSheet("font-size: 16px; font-weight: 700; color: #1E293B; margin-bottom: 5px;")
        self.settings_layout.addWidget(print_title)
        
        r_pr = QHBoxLayout(); self.cb_pr = QComboBox(); self.cb_pr.addItem("ğŸ–¥ï¸ é»˜è®¤æ‰“å°æœº/PDF")
        if platform.system() in ["Windows", "Linux"]: 
            for p in QPrinterInfo.availablePrinterNames(): self.cb_pr.addItem(f"ğŸ–¨ï¸ {p}")
        self.cb_pr.currentIndexChanged.connect(self.on_printer_changed)
        
        self.btn_prop = QPushButton(); self.btn_prop.setObjectName("PropBtn"); self.btn_prop.setFixedSize(32, 32)
        self.btn_prop.setIcon(Icons.get("settings", "#64748B")); self.btn_prop.setIconSize(QSize(18, 18))
        self.btn_prop.setToolTip("æ‰“å°æœºå±æ€§")
        self.btn_prop.clicked.connect(self.open_printer_props)
        self.btn_prop.setEnabled(False)
        r_pr.addWidget(self.cb_pr, 1); r_pr.addWidget(self.btn_prop); self.settings_layout.addLayout(r_pr)

        r_cp = QHBoxLayout(); self.sp_cpy = QSpinBox(); self.sp_cpy.setRange(1,99); self.sp_cpy.setSuffix(" ä»½")
        self.cb_pap = QComboBox(); self.cb_pap.addItems(["A4", "A5", "B5"]); self.cb_pap.currentTextChanged.connect(self.show_layout_preview)
        r_cp.addWidget(QLabel("ä»½æ•°:")); r_cp.addWidget(self.sp_cpy); r_cp.addWidget(QLabel("çº¸å¼ :")); r_cp.addWidget(self.cb_pap); self.settings_layout.addLayout(r_cp)
        
        self.settings_layout.addWidget(QLabel("æ’ç‰ˆæ¨¡å¼:"))
        rm = QHBoxLayout(); 
        self.b1 = QToolButton(); self.b1.setObjectName("LayoutCard"); self.b1.setFixedSize(85, 85); self.b1.setIconSize(QSize(72,72))
        self.b2 = QToolButton(); self.b2.setObjectName("LayoutCard"); self.b2.setFixedSize(85, 85); self.b2.setIconSize(QSize(72,72))
        self.b4 = QToolButton(); self.b4.setObjectName("LayoutCard"); self.b4.setFixedSize(85, 85); self.b4.setIconSize(QSize(72,72))
        self.b1.setCheckable(True); self.b2.setCheckable(True); self.b4.setCheckable(True)
        grp=QButtonGroup(self); grp.addButton(self.b1); grp.addButton(self.b2); grp.addButton(self.b4); self.b1.setChecked(True)
        grp.buttonClicked.connect(self.show_layout_preview)
        rm.addWidget(self.b1); rm.addWidget(self.b2); rm.addWidget(self.b4); self.settings_layout.addLayout(rm)

        r_dir = QHBoxLayout()
        self.rd_p = QRadioButton("çºµå‘"); self.rd_l = QRadioButton("æ¨ªå‘"); self.rd_l.setChecked(True)
        self.rd_p.toggled.connect(self.update_layout_icons); self.rd_l.toggled.connect(self.update_layout_icons)
        r_dir.addWidget(QLabel("æ–¹å‘:")); r_dir.addWidget(self.rd_p); r_dir.addWidget(self.rd_l); r_dir.addStretch()
        self.settings_layout.addLayout(r_dir)
        
        r_opt = QHBoxLayout()
        self.chk_cut = QCheckBox("æ˜¾ç¤ºè£å‰ªè¾…åŠ©çº¿"); self.chk_cut.setChecked(True); self.chk_cut.stateChanged.connect(self.show_layout_preview)
        self.chk_rotate = QCheckBox("å¼ºåŠ›æ‰“å°çº å"); self.chk_rotate.setToolTip("å¼ºåˆ¶æ—‹è½¬90åº¦æ‰“å°ï¼Œè§£å†³éƒ¨åˆ†æ‰“å°æœºæ–¹å‘é”™è¯¯é—®é¢˜")
        r_opt.addWidget(self.chk_cut); r_opt.addSpacing(20); r_opt.addWidget(self.chk_rotate); r_opt.addStretch()
        self.settings_layout.addLayout(r_opt)
        
        rv.addWidget(self.settings_card)

        c3 = Card(); l3 = QVBoxLayout(c3); l3.setSpacing(10); l3.setContentsMargins(20,20,20,20)
        self.lbl_inf = QLabel("0 å¼ å‘ç¥¨"); self.lbl_tot = QLabel("Â¥ 0.00", styleSheet="font-size:22px; font-weight:bold; color:#007AFF")
        l3.addWidget(self.lbl_inf); l3.addWidget(self.lbl_tot)
        self.btn_xls = QPushButton(" å¯¼å‡º Excel"); self.btn_xls.setIcon(Icons.get("excel")); self.btn_xls.clicked.connect(self.xls); l3.addWidget(self.btn_xls); rv.addWidget(c3)

        self.btn_go = QPushButton(" å¼€å§‹æ‰“å°"); self.btn_go.setObjectName("PrimaryBtn"); self.btn_go.setIcon(Icons.get("print", "white")); self.btn_go.setMinimumHeight(50)
        self.btn_go.clicked.connect(self.run); rv.addWidget(self.btn_go)
        
        btn_about = QPushButton(" å…³äºæœ¬è½¯ä»¶"); btn_about.clicked.connect(lambda: AboutDialog(self).exec())
        rv.addWidget(btn_about); rv.addStretch()

        layout.addWidget(left); layout.addWidget(mid, 1); layout.addWidget(self.right_panel)
        self.drag.upd("#555")

    def change_theme(self, mode):
        self.theme_c = ThemeManager.apply(QApplication.instance(), mode)
        self.drag.upd(self.theme_c)
        self.btn_set.setIcon(Icons.get("settings", self.theme_c))
        self.btn_del.setIcon(Icons.get("trash", "#d73a49")) 
        self.btn_xls.setIcon(Icons.get("excel", self.theme_c))
        self.btn_go.setIcon(Icons.get("print", "white"))
        self.update_layout_icons() 

    def update_layout_icons(self):
        if not hasattr(self, 'rd_l') or not self.rd_l: return
        try: is_l = self.rd_l.isChecked()
        except RuntimeError: return 
        
        icon_1 = "icon_1x1_l.png" if is_l else "icon_1x1_p.png"
        if os.path.exists(resource_path(icon_1)): self.b1.setIcon(QIcon(resource_path(icon_1)))
        else: self.b1.setIcon(Icons.get("layout_1x1_card", self.theme_c))
        icon_2 = "icon_1x2_l.png" if is_l else "icon_1x2_p.png"
        if os.path.exists(resource_path(icon_2)): self.b2.setIcon(QIcon(resource_path(icon_2)))
        else: self.b2.setIcon(Icons.get("layout_1x2_card_h" if is_l else "layout_1x2_card_v", self.theme_c))
        icon_4 = "icon_2x2_l.png" if is_l else "icon_2x2_p.png"
        if os.path.exists(resource_path(icon_4)): self.b4.setIcon(QIcon(resource_path(icon_4)))
        else: self.b4.setIcon(Icons.get("layout_2x2_card", self.theme_c))
        btn_size = QSize(90, 65) if is_l else QSize(65, 90)
        icon_size = QSize(86, 61) if is_l else QSize(61, 86)
        for btn in [self.b1, self.b2, self.b4]:
            btn.setFixedSize(btn_size)
            btn.setIconSize(icon_size)
        self.show_layout_preview()
        
        # è‡ªåŠ¨æ¸…ç†å¯èƒ½äº§ç”Ÿçš„é”™è¯¯ç©ºæ•°æ®ï¼ˆä¹‹å‰ç‰ˆæœ¬çš„bugï¼‰
        get_db().delete_invoice("")

    def _save_d_to_db(self, d):
        """è¾…åŠ©æ–¹æ³•ï¼šå°†UIæ•°æ®å­—å…¸è½¬æ¢ä¸ºæ•°æ®åº“æ ¼å¼å¹¶ä¿å­˜"""
        # åŸºç¡€æ•°æ®æ¥è‡ª ext (OCR/è§£æç»“æœ)
        info = d.get("ext", {}).copy()
        
        # å¼ºåˆ¶è¦†ç›–æ ¸å¿ƒå­—æ®µï¼ˆä»¥UIæ˜¾ç¤ºçš„ä¸ºå‡†ï¼Œæ”¯æŒç”¨æˆ·æ‰‹åŠ¨ä¿®æ”¹ï¼‰
        info["file_path"] = d.get("p")
        info["file_name"] = d.get("n")
        info["amount"] = d.get("a", 0.0)
        info["date"] = d.get("d", "")
        
        # ä¿å­˜åˆ°æ•°æ®åº“
        get_db().save_invoice(info)

    def delete_specific_item(self, item):
        row = self.list.row(item); self.list.takeItem(row); 
        d = self.data.pop(row)
        get_db().delete_invoice(d['p'])
        self.calc(); self.show_layout_preview()
    
    def show_single_doc(self, item):
        row = self.list.row(item)
        if row < len(self.data):
            f = self.data[row]['p']
            o = "H" if self.rd_l.isChecked() else "V"
            paper = self.cb_pap.currentText().replace("çº¸å¼ : ", "") if "çº¸å¼ : " in self.cb_pap.currentText() else self.cb_pap.currentText()
            doc = PDFEngine.merge([f], "1x1", paper, o, self.chk_cut.isChecked(), out_path=None)
            if doc:
                # ä½¿ç”¨å¹³å°è‡ªé€‚åº”çš„æ¸²æŸ“åˆ†è¾¨ç‡
                scale = UI_CONFIG.get("preview_render_scale", 4.0)
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(scale, scale))
                img = QImage.fromData(pix.tobytes("ppm"))
                
                # [V3.4.0] å•å¼ é¢„è§ˆä¹Ÿéœ€è¦åå‘æ—‹è½¬ä¿®å¤ (ä¸æ’ç‰ˆé¢„è§ˆé€»è¾‘ä¿æŒä¸€è‡´)
                if o == "H":
                    transform = QTransform()
                    transform.rotate(-90) # ä¿®æ­£: é€†æ—¶é’ˆ90åº¦
                    img = img.transformed(transform)
                    
                self.single_viewer.set_image(img) 
                doc.close() 
            self.stack.setCurrentIndex(1)

    def show_layout_preview(self): self.stack.setCurrentIndex(0); self.trigger_refresh()
    def edit_item(self, item):
        row = self.list.row(item)
        old_val = self.data[row].get('a', 0)
        val, ok = QInputDialog.getDouble(self, "ä¿®æ­£é‡‘é¢", "è¯·è¾“å…¥æ­£ç¡®é‡‘é¢:", old_val, 0.00, 1000000, 2)
        if ok:
            self.data[row]['a'] = val
            self.data[row]['manually_edited'] = True
            self._save_d_to_db(self.data[row])
            widget = self.list.itemWidget(item)
            if widget:
                widget.update_display(self.data[row])
            self.calc()
    
    def on_printer_changed(self, idx):
        if idx == 0:
            self.btn_prop.setEnabled(False)
            self.btn_go.setText(" é¢„è§ˆ / ç”ŸæˆPDF")
        else:
            self.btn_prop.setEnabled(True)
            p_name = self.cb_pr.currentText().replace("ğŸ–¨ï¸ ", "")
            self.current_printer = QPrinter(QPrinterInfo.printerInfo(p_name), QPrinter.PrinterMode.HighResolution)
            self.btn_go.setText(f" æ‰“å°åˆ°: {p_name[:8]}...")

    def open_printer_props(self):
        dlg = QPrintDialog(self.current_printer, self)
        if dlg.exec() == QDialog.DialogCode.Accepted: pass 

    def trigger_refresh(self): self.preview_timer.start(200)
    def generate_realtime_preview(self):
        m="1x1"; m="1x2" if self.b2.isChecked() else m; m="2x2" if self.b4.isChecked() else m
        o="H" if self.rd_l.isChecked() else "V"; 
        
        rotate_preview = (o == "H")

        if hasattr(self, 'current_doc') and self.current_doc: self.current_doc.close(); self.current_doc = None
        gc.collect()
        paper = self.cb_pap.currentText().replace("çº¸å¼ : ", "") if "çº¸å¼ : " in self.cb_pap.currentText() else self.cb_pap.currentText()
        self.current_doc = PDFEngine.merge([x['p'] for x in self.data], m, paper, o, self.chk_cut.isChecked(), out_path=None)
        page_imgs = []
        if self.current_doc:
            for page in self.current_doc: 
                # ä½¿ç”¨å¹³å°è‡ªé€‚åº”çš„æ¸²æŸ“åˆ†è¾¨ç‡
                scale = UI_CONFIG.get("preview_render_scale", 4.0)
                pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale)); img = QImage.fromData(pix.tobytes("ppm"))
                if rotate_preview:
                    transform = QTransform()
                    transform.rotate(-90) 
                    img = img.transformed(transform)
                page_imgs.append(img)
        self.word_preview.show_pages(page_imgs)

    def add_files(self, fs):
        """æ·»åŠ æ–‡ä»¶å¹¶å¼‚æ­¥æ‰§è¡Œ OCR è¯†åˆ«"""
        logger = logging.getLogger(__name__)
        logger.info(f"å¼€å§‹æ·»åŠ  {len(fs)} ä¸ªæ–‡ä»¶")
        
        try:
            s = QSettings("MySoft", "InvoiceMaster")
            ak, sk = s.value("ak"), s.value("sk")
            
            # æ˜¾ç¤ºå¯¼å…¥è¿›åº¦å¯¹è¯æ¡†
            import_progress = ProgressDialog(self, "å‘ç¥¨å¯¼å…¥ä¸­", can_cancel=False)
            import_progress.show()
            QApplication.processEvents()
            
            # å…ˆåŒæ­¥æ·»åŠ æ‰€æœ‰æ–‡ä»¶åˆ°åˆ—è¡¨ï¼ˆå¿«é€Ÿå“åº”ç”¨æˆ·ï¼‰
            files_with_index = []
            start_idx = len(self.data)
            
            added_count = 0
            total_files = len(fs)
            
            for i, f in enumerate(fs):
                try:
                    # [V3.5] æ–‡ä»¶åé¢„è¿‡æ»¤ï¼šç›´æ¥è·³è¿‡æ˜ç¡®çš„éå‘ç¥¨æ–‡ä»¶
                    # åŒ…å«ï¼šæ¸…å•ã€å…¥ä½å‡­è¯ã€è¡Œç¨‹æŠ¥é”€å•ã€ç»“ç®—å•
                    logger.info(f"æ·»åŠ æ–‡ä»¶: {os.path.basename(f)}")
                    # [V3.5] ç§»é™¤æ–‡ä»¶åè¿‡æ»¤ï¼Œå…è®¸æ‰€æœ‰æ–‡ä»¶å¯¼å…¥ï¼Œä½†åœ¨ç»Ÿè®¡æ—¶æ’é™¤
                    basename = os.path.basename(f)
                    # ignore_keywords = ["æ¸…å•", "å…¥ä½å‡­è¯", "è¡Œç¨‹æŠ¥é”€å•", "ç»“ç®—å•"]
                    # if any(k in basename for k in ignore_keywords):
                    #     logger.info(f"ğŸš« æ ¹æ®æ–‡ä»¶åè·³è¿‡: {basename}")
                    #     continue

                    logger.info(f"æ·»åŠ æ–‡ä»¶: {basename}")
                    
                    # æ›´æ–°å¯¼å…¥è¿›åº¦
                    import_progress.update_progress(i + 1, total_files, basename)
                    
                    d = {"p": f, "n": basename, "d": "", "a": 0.0, "ext": {}, "_pending_ocr": True}
                    
                    # æœ¬åœ°è§£æå®Œæ•´å‘ç¥¨ä¿¡æ¯
                    # æœ¬åœ°è§£æå®Œæ•´å‘ç¥¨ä¿¡æ¯
                    if f.lower().endswith(".pdf"):
                        local_result = InvoiceHelper.parse_invoice_local(f)
                        
                        # [V3.5] ç‰¹æ®Šå¤„ç†ï¼šå¦‚æœæ˜¯æ¸…å•æˆ–éå‘ç¥¨å‡­è¯ï¼Œç›´æ¥è®¤å¯ï¼Œè·³è¿‡OCR
                        inv_type = local_result.get("invoice_type", "") if local_result else ""
                        is_special_doc = "æ¸…å•" in inv_type or "éå‘ç¥¨" in inv_type
                        
                        if is_special_doc:
                             d["a"] = local_result.get("amount", 0)
                             d["d"] = local_result.get("date", "")
                             d["ext"] = local_result
                             d["_pending_ocr"] = False # æ˜ç¡®æ ‡è®°ä¸éœ€OCR
                             logger.info(f"âœ… è¯†åˆ«ä¸ºç‰¹æ®Šæ–‡æ¡£(ä¸è®¡å…¥ç»Ÿè®¡): {os.path.basename(f)} - {inv_type}")
                             
                        # åªæœ‰å½“æ•°æ®å®Œæ•´æ—¶æ‰è·³è¿‡OCR
                        elif InvoiceHelper._is_result_complete(local_result):
                            d["a"] = local_result["amount"]
                            d["d"] = local_result.get("date", "")
                            d["ext"] = local_result
                            d["_pending_ocr"] = False  # æ•°æ®å®Œæ•´ï¼Œä¸éœ€è¦OCR
                            logger.info(f"âœ… æœ¬åœ°è§£æå®Œæ•´: {os.path.basename(f)}, é‡‘é¢: {d['a']}")
                        elif local_result.get("amount", 0) > 0:
                            # æœ‰é‡‘é¢ä½†ä¸å®Œæ•´ï¼Œéœ€è¦ç»§ç»­OCR
                            d["a"] = local_result["amount"]
                            d["d"] = local_result.get("date", "")
                            d["ext"] = local_result
                            d["_pending_ocr"] = True  # æ•°æ®ä¸å®Œæ•´ï¼Œéœ€è¦OCRè¡¥å……
                            logger.info(f"âš ï¸ æœ¬åœ°è§£æä¸å®Œæ•´: {os.path.basename(f)}, é‡‘é¢: {d['a']}ï¼Œå°†ä½¿ç”¨OCR")
                        else:
                            logger.info(f"âš ï¸ æœ¬åœ°è§£æå¤±è´¥: {os.path.basename(f)}ï¼Œå°†ä½¿ç”¨OCR")
                    
                    item = QListWidgetItem(self.list)
                    item.setSizeHint(QSize(250, 60))
                    widget = InvoiceItemWidget(d, item, self.delete_specific_item)
                    self.list.setItemWidget(item, widget)
                    
                    
                    # [V3.5] è´¢åŠ¡ä¸¥è°¨æ€§è¿‡æ»¤ï¼šè™½ç„¶å¯¼å…¥æ¸…å•/å‡­è¯ï¼Œä½†ä¸è®¡å…¥ç»Ÿè®¡
                    # ä¹‹å‰çš„ç‰ˆæœ¬æ˜¯ç›´æ¥è·³è¿‡(continue)ï¼Œç°åœ¨æ”¹ä¸ºå¯¼å…¥ä½†æ ‡è®°ç±»å‹
                    skipped_types = ["å‘ç¥¨æ¸…å•", "éå‘ç¥¨å‡­è¯"]
                    base_invoicetype = d.get("invoice_type", "")
                    if base_invoicetype in skipped_types or "éå‘ç¥¨" in base_invoicetype:
                         # å¯ä»¥åœ¨è¿™é‡Œåšä¸€äº›é¢å¤–çš„UIæ ‡è®°ï¼Œç›®å‰ä»…ä¾é  calc() æ’é™¤ç»Ÿè®¡
                         pass
                    
                    self.data.append(d)
                    widget.update_display(d)
                    self._save_d_to_db(d)
                    added_count += 1
                    
                    # è®°å½•éœ€è¦ OCR çš„æ–‡ä»¶ï¼ˆåªæœ‰ _pending_ocr=True çš„æ‰éœ€è¦ï¼‰
                    # å¦‚æœæœ‰ç™¾åº¦API Key æˆ– ç§æœ‰OCRåœ°å€ï¼Œå°±æ·»åŠ åˆ°å¾…è¯†åˆ«åˆ—è¡¨
                    private_ocr_url = QSettings("MySoft", "InvoiceMaster").value("private_ocr_url", "")
                    if (ak or private_ocr_url) and d.get("_pending_ocr", True):
                        files_with_index.append((start_idx + i, f))
                
                except Exception as inner_e:
                    logger.error(f"å¤„ç†å•ä¸ªæ–‡ä»¶å¤±è´¥ {f}: {str(inner_e)}", exc_info=True)
                    continue

            # å…³é—­å¯¼å…¥è¿›åº¦å¯¹è¯æ¡†
            import_progress.close()
            
            self.calc()
            self.show_layout_preview()
            QApplication.processEvents()
            
            # å¦‚æœæœ‰éœ€è¦è¯†åˆ«çš„æ–‡ä»¶ï¼Œå¯åŠ¨å¼‚æ­¥ OCR
            if files_with_index:
                self._start_async_ocr(files_with_index, ak, sk)
            else:
                logger.info(f"æ–‡ä»¶æ·»åŠ å®Œæˆï¼Œå…± {len(self.data)} ä¸ªå‘ç¥¨ï¼ˆæ—  OCRï¼‰")
                
        except Exception as e:
            logger.error(f"æ·»åŠ æ–‡ä»¶å…¨å±€é”™è¯¯: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "å¯¼å…¥å¤±è´¥", f"æ·»åŠ æ–‡ä»¶æ—¶å‘ç”Ÿé”™è¯¯:\n{str(e)}")
    
    def _start_async_ocr(self, files_with_index, ak, sk):
        """å¯åŠ¨å¼‚æ­¥ OCR å¤„ç†"""
        logger = logging.getLogger(__name__)
        logger.info(f"å¯åŠ¨å¼‚æ­¥ OCRï¼Œå…± {len(files_with_index)} ä¸ªæ–‡ä»¶")
        
        # åˆ›å»ºè¿›åº¦å¯¹è¯æ¡†
        self.progress_dialog = ProgressDialog(self, "OCR è¯†åˆ«ä¸­", can_cancel=True)
        
        # åˆ›å»º OCR å·¥ä½œçº¿ç¨‹
        self.ocr_worker = OcrWorker(files_with_index, ak, sk, self)
        self.ocr_worker.progress.connect(self._on_ocr_progress)
        self.ocr_worker.result.connect(self._on_ocr_result)
        self.ocr_worker.error.connect(self._on_ocr_error)
        self.ocr_worker.finished_all.connect(self._on_ocr_finished)
        
        # å–æ¶ˆå¤„ç†
        self.progress_dialog.cancelled.connect(self.ocr_worker.cancel)
        
        # å¯åŠ¨å·¥ä½œçº¿ç¨‹å’Œæ˜¾ç¤ºå¯¹è¯æ¡†
        self.ocr_worker.start()
        self.progress_dialog.show()
    
    def _on_ocr_progress(self, current, total, filename):
        """OCR è¿›åº¦æ›´æ–°"""
        if self.progress_dialog:
            self.progress_dialog.update_progress(current, total, filename)
    
    def _on_ocr_result(self, idx, result):
        """OCR å•ä¸ªç»“æœè¿”å›"""
        logger = logging.getLogger(__name__)
        if idx < len(self.data):
            d = self.data[idx]
            d["_pending_ocr"] = False
            
            if result:
                if "amount" in result:
                    d["a"] = result["amount"]
                if "date" in result:
                    d["d"] = result["date"]
                d["ext"] = result
                logger.info(f"OCR ç»“æœå·²æ›´æ–°: {d['n']}, é‡‘é¢: {d.get('a', 0)}")
                self._save_d_to_db(d)

            
            # æ›´æ–°åˆ—è¡¨é¡¹æ˜¾ç¤º
            item = self.list.item(idx)
            if item:
                widget = self.list.itemWidget(item)
                if widget:
                    widget.update_display(d)
            
            # æ›´æ–°ç»Ÿè®¡
            self.calc()
    
    def _on_ocr_error(self, idx, error_msg):
        """OCR å•ä¸ªé”™è¯¯å¤„ç†"""
        logger = logging.getLogger(__name__)
        if idx < len(self.data):
            d = self.data[idx]
            d["_pending_ocr"] = False
            logger.warning(f"OCR å¤±è´¥ [{d['n']}]: {error_msg}")
    
    def _on_ocr_finished(self):
        """OCR å…¨éƒ¨å®Œæˆ"""
        logger = logging.getLogger(__name__)
        logger.info(f"å¼‚æ­¥ OCR å¤„ç†å®Œæˆï¼Œå…± {len(self.data)} ä¸ªå‘ç¥¨")
        
        if self.progress_dialog:
            self.progress_dialog.close()
            self.progress_dialog = None
        
        self.ocr_worker = None
        self.calc()
        self.show_layout_preview()

    def calc(self):
        """è®¡ç®—ç»Ÿè®¡ä¿¡æ¯ï¼ˆä»…è®¡ç®—æœ‰æ•ˆå‘ç¥¨ï¼Œæ’é™¤æ¸…å•å’Œéå‘ç¥¨å‡­è¯ï¼‰"""
        total_n = 0
        total_a = 0.0
        unrecognized_n = 0  # æœªè¯†åˆ«æ•°é‡
        
        ignored_types = ["å‘ç¥¨æ¸…å•", "éå‘ç¥¨å‡­è¯"]
        
        for d in self.data:
            # [ä¿®å¤] ä» ext ä¸­è¯»å– invoice_typeï¼ˆè€Œé d ç›´æ¥è¯»å–ï¼‰
            inv_type = d.get("ext", {}).get("invoice_type", "")
            # æ’é™¤æ¸…å•å’Œéå‘ç¥¨å‡­è¯
            if inv_type in ignored_types or "éå‘ç¥¨" in inv_type:
                continue
            
            # ç»Ÿè®¡æœªè¯†åˆ«å‘ç¥¨ï¼ˆé‡‘é¢ä¸º0æˆ–æ— æ—¥æœŸï¼‰
            amount = d.get("a", 0)
            date = d.get("d", "")
            if amount == 0 or not date:
                unrecognized_n += 1
            
            total_n += 1
            total_a += amount
        
        # æ˜¾ç¤ºæ ¼å¼ï¼šå·²è¯†åˆ«æ•°é‡ + æœªè¯†åˆ«æ•°é‡
        if unrecognized_n > 0:
            self.lbl_inf.setText(f"{total_n} å¼ å‘ç¥¨ï¼Œ{unrecognized_n} å¼ æœªè¯†åˆ«")
        else:
            self.lbl_inf.setText(f"{total_n} å¼ å‘ç¥¨")
        self.lbl_tot.setText(f"Â¥ {total_a:,.2f}")
    def clear(self): self.list.clear(); self.data=[]; self.calc(); self.trigger_refresh()
    def ctx_menu(self, p): m=QMenu(); a=QAction("åˆ é™¤",self); a.triggered.connect(self.del_sel); m.addAction(a); m.exec(self.list.mapToGlobal(p))
    def del_sel(self):
        for r in sorted([self.list.row(i) for i in self.list.selectedItems()], reverse=True): 
            self.list.takeItem(r); 
            d = self.data.pop(r)
            get_db().delete_invoice(d['p'])
        self.calc(); self.trigger_refresh()
    def xls(self):
        logger = logging.getLogger(__name__)
        if not self.data: return
        
        # æ£€æŸ¥æ¿€æ´»çŠ¶æ€
        info = self.license_manager.get_activation_info()
        if not info['is_activated']:
            if info['remaining_trials'] <= 0:
                # è¯•ç”¨æ¬¡æ•°ç”¨å®Œï¼Œå¿…é¡»æ¿€æ´»
                QMessageBox.warning(self, "éœ€è¦æ¿€æ´»", "è¯•ç”¨æ¬¡æ•°å·²ç”¨å®Œï¼Œè¯·æ¿€æ´»è½¯ä»¶åç»§ç»­ä½¿ç”¨ã€‚")
                dialog = ActivationDialog(self, self.license_manager)
                if dialog.exec() != QDialog.DialogCode.Accepted:
                    return
                # æ¿€æ´»æˆåŠŸåç»§ç»­
            else:
                # è¿˜æœ‰è¯•ç”¨æ¬¡æ•°ï¼Œæ˜¾ç¤ºæç¤ºå¹¶è¯¢é—®æ˜¯å¦ç»§ç»­
                remaining = info['remaining_trials']
                reply = QMessageBox.question(
                    self, 
                    "è¯•ç”¨æç¤º", 
                    f"æ‚¨è¿˜æœ‰ {remaining} æ¬¡å…è´¹å¯¼å‡ºæœºä¼šã€‚\n\næ˜¯å¦ç»§ç»­å¯¼å‡ºï¼Ÿ\nï¼ˆå¯¼å‡ºæˆåŠŸåå°†ä½¿ç”¨1æ¬¡æœºä¼šï¼‰",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return  # ç”¨æˆ·é€‰æ‹©ä¸ç»§ç»­ï¼Œç›´æ¥è¿”å›ï¼Œä¸æ‰£æ¬¡æ•°
        
        logger.info(f"å¼€å§‹å¯¼å‡º Excel: {len(self.data)} æ¡æ•°æ®")
        
        # è¯»å–ä¸Šæ¬¡ä¿å­˜çš„è·¯å¾„
        s = QSettings("MySoft", "InvoiceMaster")
        last_path = s.value("last_excel_path", os.path.expanduser("~/Desktop/invoice_report.xlsx"))
        
        # å…ˆæ£€æŸ¥é»˜è®¤è·¯å¾„æ–‡ä»¶æ˜¯å¦å­˜åœ¨ï¼Œæä¾›ä¸­æ–‡é€‰é¡¹
        file_action = "new"  # new=æ–°å»º, append=è¿½åŠ , overwrite=è¦†ç›–
        
        if os.path.exists(last_path):
            # æ–‡ä»¶å·²å­˜åœ¨ï¼Œæ˜¾ç¤ºä¸­æ–‡é€‰æ‹©å¯¹è¯æ¡†
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("æ–‡ä»¶å·²å­˜åœ¨")
            msg_box.setText(f"æ–‡ä»¶ \"{os.path.basename(last_path)}\" å·²å­˜åœ¨ã€‚\n\nè¯·é€‰æ‹©æ“ä½œæ–¹å¼ï¼š")
            msg_box.setIcon(QMessageBox.Icon.Question)
            
            append_btn = msg_box.addButton("ğŸ“¥ è¿½åŠ æ•°æ®", QMessageBox.ButtonRole.AcceptRole)
            overwrite_btn = msg_box.addButton("ğŸ”„ è¦†ç›–æ–‡ä»¶", QMessageBox.ButtonRole.DestructiveRole)
            newfile_btn = msg_box.addButton("ğŸ“ å¦å­˜ä¸º...", QMessageBox.ButtonRole.ActionRole)
            cancel_btn = msg_box.addButton("å–æ¶ˆ", QMessageBox.ButtonRole.RejectRole)
            
            msg_box.exec()
            clicked = msg_box.clickedButton()
            
            if clicked == cancel_btn:
                return
            elif clicked == append_btn:
                file_action = "append"
                p = last_path
            elif clicked == overwrite_btn:
                file_action = "overwrite"
                p = last_path
            elif clicked == newfile_btn:
                # ç”¨æˆ·é€‰æ‹©å¦å­˜ä¸ºï¼Œæ‰“å¼€æ–‡ä»¶å¯¹è¯æ¡†
                p, _ = QFileDialog.getSaveFileName(self, "ä¿å­˜ Excel æŠ¥è¡¨", last_path, "Excel (*.xlsx)",
                                                   options=QFileDialog.Option.DontConfirmOverwrite)
                if not p: return
                file_action = "overwrite" if os.path.exists(p) else "new"
        else:
            # æ–‡ä»¶ä¸å­˜åœ¨ï¼Œä½¿ç”¨é»˜è®¤è·¯å¾„æˆ–è®©ç”¨æˆ·é€‰æ‹©
            p, _ = QFileDialog.getSaveFileName(self, "ä¿å­˜ Excel æŠ¥è¡¨", last_path, "Excel (*.xlsx)",
                                               options=QFileDialog.Option.DontConfirmOverwrite)
            if not p: return
            
            # æ£€æŸ¥ç”¨æˆ·é€‰æ‹©çš„æ–°è·¯å¾„æ˜¯å¦å­˜åœ¨
            if os.path.exists(p):
                reply = QMessageBox.question(
                    self, "ç¡®è®¤è¦†ç›–",
                    f"æ–‡ä»¶ \"{os.path.basename(p)}\" å·²å­˜åœ¨ã€‚\n\næ˜¯å¦è¦†ç›–è¯¥æ–‡ä»¶ï¼Ÿ",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.No
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return
                file_action = "overwrite"
        
        # ä¿å­˜è·¯å¾„ä¾›ä¸‹æ¬¡ä½¿ç”¨
        s.setValue("last_excel_path", p)
        
        try:
            # å‡†å¤‡æ–°æ•°æ® - 18ä¸ªä¸“ä¸šå­—æ®µ
            new_rows = []
            from datetime import datetime
            import_time = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            export_idx = 1
            for x in self.data:
                ext = x.get("ext", {})
                
                # [V3.5] å¯¼å‡ºè¿‡æ»¤ï¼šæ¸…å•å’Œéå‘ç¥¨å‡­è¯ä¸å¯¼å‡ºåˆ°Excel
                # è¿™äº›åªæ˜¯ä¸ºäº†ç®¡ç†æŸ¥çœ‹ï¼Œä¸åº”è¿›å…¥è´¢åŠ¡æŠ¥è¡¨
                inv_type = ext.get("invoice_type", "")
                skipped_types = ["å‘ç¥¨æ¸…å•", "éå‘ç¥¨å‡­è¯"]
                if inv_type in skipped_types or "éå‘ç¥¨" in inv_type:
                    continue
                
                # å¤„ç†é‡‘é¢å­—æ®µ,ç¡®ä¿æ˜¯æ•°å€¼ç±»å‹
                try: amount = float(x.get("a", 0) or 0)
                except: amount = 0
                try: amount_without_tax = float(ext.get("amount_without_tax", "") or 0)
                except: amount_without_tax = ""
                try: tax_amt = float(ext.get("tax_amt", "") or 0)
                except: tax_amt = ""
                
                # æ ‡å‡†å¯¼å‡ºå­—æ®µ
                row_data = {
                    "åºå·": export_idx,
                    "å¼€ç¥¨æ—¥æœŸ": x.get("d", ""), 
                    "å‘ç¥¨ç±»å‹": ext.get("invoice_type", ""),
                    "å‘ç¥¨ä»£ç ": ext.get("code", ""), 
                    "å‘ç¥¨å·ç ": ext.get("number", ""), 
                    "æ ¡éªŒç ": ext.get("check_code", "")[-6:] if ext.get("check_code") else "",
                    "è´­ä¹°æ–¹åç§°": ext.get("buyer", ""),
                    "è´­ä¹°æ–¹ç¨å·": ext.get("buyer_tax_id", ""),
                    "é”€å”®æ–¹åç§°": ext.get("seller", ""), 
                    "é”€å”®æ–¹ç¨å·": ext.get("seller_tax_id", ""),
                    "ä¸å«ç¨é‡‘é¢": amount_without_tax,
                    "ç¨ç‡": ext.get("tax_rate", ""),
                    "ç¨é¢": tax_amt,
                    "ä»·ç¨åˆè®¡": amount,
                    "å•†å“æ˜ç»†": ext.get("item_name", ""),
                    "å¤‡æ³¨": ext.get("remark", ""),
                    "å¯¼å…¥æ—¶é—´": import_time,
                    "æ–‡ä»¶è·¯å¾„": x.get("p", "")
                }
                
                new_rows.append(row_data)
                export_idx += 1
            
            new_df = pd.DataFrame(new_rows)
            # ç¡®ä¿åˆ—é¡ºåº
            fields = ["åºå·", "å¼€ç¥¨æ—¥æœŸ", "å‘ç¥¨ç±»å‹", "å‘ç¥¨ä»£ç ", "å‘ç¥¨å·ç ", "æ ¡éªŒç ", 
                     "è´­ä¹°æ–¹åç§°", "è´­ä¹°æ–¹ç¨å·", "é”€å”®æ–¹åç§°", "é”€å”®æ–¹ç¨å·", 
                     "ä¸å«ç¨é‡‘é¢", "ç¨ç‡", "ç¨é¢", "ä»·ç¨åˆè®¡", "å•†å“æ˜ç»†", "å¤‡æ³¨", "å¯¼å…¥æ—¶é—´", "æ–‡ä»¶è·¯å¾„"]
            new_df = new_df[fields]
            
            # æ ¹æ®ç”¨æˆ·é€‰æ‹©å†³å®šæ˜¯è¿½åŠ è¿˜æ˜¯è¦†ç›–
            if file_action == "append" and os.path.exists(p):
                try:
                    existing_df = pd.read_excel(p)
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                    is_append = True
                except Exception:
                    combined_df = new_df
                    is_append = False
            else:
                combined_df = new_df
                is_append = False
            
            # ä¿å­˜åˆ° Excel
            combined_df.to_excel(p, index=False, engine='openpyxl')
            
            # ä½¿ç”¨ openpyxl æ·»åŠ ä¸“ä¸šæ ·å¼
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                wb = load_workbook(p)
                ws = wb.active
                
                # å®šä¹‰æ ·å¼
                header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin', color='E2E8F0'),
                    right=Side(style='thin', color='E2E8F0'),
                    top=Side(style='thin', color='E2E8F0'),
                    bottom=Side(style='thin', color='E2E8F0')
                )
                
                # åº”ç”¨è¡¨å¤´æ ·å¼
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin_border
                
                # å†»ç»“é¦–è¡Œ
                ws.freeze_panes = "A2"
                
                # å»ºç«‹å­—æ®µåˆ°åˆ—ç´¢å¼•çš„æ˜ å°„ (1-based)
                field_list = ["åºå·", "å¼€ç¥¨æ—¥æœŸ", "å‘ç¥¨ç±»å‹", "å‘ç¥¨ä»£ç ", "å‘ç¥¨å·ç ", "æ ¡éªŒç ", 
                             "è´­ä¹°æ–¹åç§°", "è´­ä¹°æ–¹ç¨å·", "é”€å”®æ–¹åç§°", "é”€å”®æ–¹ç¨å·", 
                             "ä¸å«ç¨é‡‘é¢", "ç¨ç‡", "ç¨é¢", "ä»·ç¨åˆè®¡", "å•†å“æ˜ç»†", "å¤‡æ³¨", "å¯¼å…¥æ—¶é—´", "æ–‡ä»¶è·¯å¾„"]
                col_map = {field: i+1 for i, field in enumerate(field_list)}
                
                # è®¾ç½®åˆ—å®½
                field_widths = {
                    "åºå·": 6, "å¼€ç¥¨æ—¥æœŸ": 12, "å‘ç¥¨ç±»å‹": 14, "å‘ç¥¨ä»£ç ": 14, "å‘ç¥¨å·ç ": 12,
                    "æ ¡éªŒç ": 10, "è´­ä¹°æ–¹åç§°": 25, "è´­ä¹°æ–¹ç¨å·": 22, "é”€å”®æ–¹åç§°": 25, "é”€å”®æ–¹ç¨å·": 22,
                    "ä¸å«ç¨é‡‘é¢": 12, "ç¨ç‡": 8, "ç¨é¢": 12, "ä»·ç¨åˆè®¡": 14, "å•†å“æ˜ç»†": 30,
                    "å¤‡æ³¨": 20, "å¯¼å…¥æ—¶é—´": 18, "æ–‡ä»¶è·¯å¾„": 40
                }
                for i, field in enumerate(field_list, 1):
                    if field in field_widths:
                        ws.column_dimensions[get_column_letter(i)].width = field_widths[field]
                
                # æ£€æµ‹é‡å¤å‘ç¥¨ï¼šå‘ç¥¨å·ç  + å¼€ç¥¨æ—¥æœŸ + é‡‘é¢ ä¸‰è€…éƒ½ç›¸åŒæ‰ç®—é‡å¤
                invoice_keys = {}
                duplicate_rows = set()
                num_col_idx = col_map.get("å‘ç¥¨å·ç ")
                date_col_idx = col_map.get("å¼€ç¥¨æ—¥æœŸ")
                amount_col_idx = col_map.get("ä»·ç¨åˆè®¡")
                
                if num_col_idx:
                    for row_idx in range(2, ws.max_row + 1):
                        invoice_num = ws.cell(row=row_idx, column=num_col_idx).value
                        invoice_date = ws.cell(row=row_idx, column=date_col_idx).value if date_col_idx else ""
                        invoice_amount = ws.cell(row=row_idx, column=amount_col_idx).value if amount_col_idx else ""
                        
                        # ç»„åˆé”®ï¼šå·ç +æ—¥æœŸ+é‡‘é¢
                        if invoice_num and str(invoice_num).strip():
                            key = f"{invoice_num}|{invoice_date}|{invoice_amount}"
                            if key in invoice_keys:
                                duplicate_rows.add(row_idx)
                                duplicate_rows.add(invoice_keys[key])
                            else:
                                invoice_keys[key] = row_idx
                
                # é‡‘é¢åˆ—ç´¢å¼•
                amount_cols = [col_map.get(f) for f in ["ä¸å«ç¨é‡‘é¢", "ç¨é¢", "ä»·ç¨åˆè®¡"] if f in col_map]
                
                # åº”ç”¨æ•°æ®è¡Œæ ·å¼(æ–‘é©¬çº¹ + é‡å¤æ ‡è®°)
                for row_idx in range(2, ws.max_row + 1):
                    is_duplicate = row_idx in duplicate_rows
                    is_zebra = row_idx % 2 == 0
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        if is_duplicate:
                            cell.fill = yellow_fill
                        elif is_zebra:
                            cell.fill = zebra_fill
                        # é‡‘é¢åˆ—å³å¯¹é½
                        if col_idx in amount_cols:
                            cell.alignment = Alignment(horizontal="right")
                
                # æ·»åŠ åº•éƒ¨ç»Ÿè®¡è¡Œ
                # è®¡ç®—æ€»é‡‘é¢
                total_amount = 0
                amount_col_idx = col_map.get("ä»·ç¨åˆè®¡")
                
                if amount_col_idx:
                    for row_idx in range(2, ws.max_row + 1):  # éå†åˆ°ç°æœ‰æœ€å¤§è¡Œ
                        try:
                            val = ws.cell(row=row_idx, column=amount_col_idx).value
                            if val: total_amount += float(val)
                        except: pass
                
                stat_row = ws.max_row + 1
                
                # åœ¨ç¬¬ä¸€åˆ—æˆ–"åºå·"åˆ—æ˜¾ç¤º"ç»Ÿè®¡"
                label_col = col_map.get("åºå·", 1)
                ws.cell(row=stat_row, column=label_col, value="ç»Ÿè®¡").font = Font(bold=True)
                
                # åœ¨ç¬¬äºŒåˆ—æˆ–"å¼€ç¥¨æ—¥æœŸ"åˆ—æ˜¾ç¤ºæ•°é‡
                count_col = col_map.get("å¼€ç¥¨æ—¥æœŸ", 2)
                ws.cell(row=stat_row, column=count_col, value=f"å…± {len(new_rows)} å¼ å‘ç¥¨")
                
                # æ˜¾ç¤ºæ€»é‡‘é¢
                if amount_col_idx:
                    ws.cell(row=stat_row, column=amount_col_idx, value=f"Â¥{total_amount:,.2f}").font = Font(bold=True, color="DC2626")
                
                # æ·»åŠ å·¥ä½œè¡¨ä¿æŠ¤(å®‰å…¨é”å®šåŠŸèƒ½)
                # å…è®¸: é€‰æ‹©å•å…ƒæ ¼ã€å¤åˆ¶ã€æ’åºã€ç­›é€‰ã€æŸ¥æ‰¾
                # ç¦æ­¢: ç¼–è¾‘å†…å®¹ã€åˆ é™¤è¡Œåˆ—ã€ä¿®æ”¹æ ¼å¼ã€æ’å…¥è¡Œåˆ—
                from openpyxl.worksheet.protection import SheetProtection
                SHEET_PASSWORD = "InvoiceMaster2024"  # ä¿æŠ¤å¯†ç 
                
                ws.protection = SheetProtection(
                    sheet=True,
                    password=SHEET_PASSWORD,
                    selectLockedCells=False,
                    selectUnlockedCells=False,
                    sort=False,
                    autoFilter=False,
                    formatCells=True,
                    formatColumns=True,
                    formatRows=True,
                    insertColumns=True,
                    insertRows=True,
                    insertHyperlinks=True,
                    deleteColumns=True,
                    deleteRows=True,
                    objects=False,
                    scenarios=False,
                    pivotTables=False,
                )
                logger.info("Excel å·¥ä½œè¡¨ä¿æŠ¤å·²å¯ç”¨")
                
                wb.save(p)
                
                # æç¤ºä¿¡æ¯
                if is_append:
                    msg = f"âœ… å·²è¿½åŠ  {len(new_df)} æ¡æ•°æ®åˆ°ç°æœ‰æ–‡ä»¶ï¼\n"
                    logger.info(f"Excel è¿½åŠ æˆåŠŸ: {p}, æ–°å¢ {len(new_df)} æ¡")
                else:
                    msg = f"âœ… å·²å¯¼å‡º {len(new_df)} æ¡æ•°æ®ï¼\n"
                    logger.info(f"Excel å¯¼å‡ºæˆåŠŸ: {p}, å…± {len(new_df)} æ¡")
                
                msg += f"ğŸ’° ä»·ç¨åˆè®¡: Â¥{total_amount:,.2f}\n"
                
                if duplicate_rows:
                    msg += f"âš ï¸ æ£€æµ‹åˆ° {len(duplicate_rows)//2} ç»„é‡å¤å‘ç¥¨ï¼ˆå·²ç”¨é»„è‰²æ ‡è®°ï¼‰\n"
                    logger.warning(f"æ£€æµ‹åˆ° {len(duplicate_rows)} æ¡é‡å¤å‘ç¥¨")
                else:
                    msg += "âœ“ æœªæ£€æµ‹åˆ°é‡å¤å‘ç¥¨\n"
                
                msg += "ğŸ”’ å·¥ä½œè¡¨å·²ä¿æŠ¤ï¼Œä»…å…è®¸æŸ¥çœ‹å’Œå¤åˆ¶"
                
                # å¯¼å‡ºæˆåŠŸåï¼Œæ‰£é™¤è¯•ç”¨æ¬¡æ•°ï¼ˆå¦‚æœæœªæ¿€æ´»ï¼‰
                info = self.license_manager.get_activation_info()
                if not info['is_activated']:
                    self.license_manager.increment_trial_count()
                    logger.info("è¯•ç”¨æ¬¡æ•°å·²æ‰£é™¤")
                
                QMessageBox.information(self, "å¯¼å‡ºæˆåŠŸ", msg)
                
            except Exception as e:
                # å¦‚æœé¢œè‰²æ ‡è®°å¤±è´¥ï¼Œè‡³å°‘æ•°æ®å·²ä¿å­˜
                logger.warning(f"Excel é¢œè‰²æ ‡è®°å¤±è´¥: {str(e)}")
                QMessageBox.information(self, "å¯¼å‡ºæˆåŠŸ", f"æ•°æ®å·²å¯¼å‡ºï¼Œä½†é¢œè‰²æ ‡è®°å¤±è´¥: {str(e)}")
                
        except Exception as e:
            logger.error(f"Excel å¯¼å‡ºå¤±è´¥: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "å¯¼å‡ºå¤±è´¥", f"é”™è¯¯: {str(e)}")
    def run(self):
        """æ‰§è¡Œæ‰“å°æ“ä½œï¼ˆå¼‚æ­¥ï¼‰"""
        if not self.data: 
            return QMessageBox.warning(self, "Tips", "è¯·å…ˆæ·»åŠ å‘ç¥¨")
        
        self.btn_go.setText("å¤„ç†ä¸­...")
        self.btn_go.setEnabled(False)
        QApplication.processEvents()
        
        # å‡†å¤‡å‚æ•°
        m = "1x1"
        m = "1x2" if self.b2.isChecked() else m
        m = "2x2" if self.b4.isChecked() else m
        o = "H" if self.rd_l.isChecked() else "V"
        out = os.path.expanduser("~/Desktop/Print_Job.pdf")
        
        if out not in self.temp_files:
            self.temp_files.append(out)
        
        paper = self.cb_pap.currentText().replace("çº¸å¼ : ", "") if "çº¸å¼ : " in self.cb_pap.currentText() else self.cb_pap.currentText()
        
        # ä¿å­˜æ‰“å°å‚æ•°ä¾›å›è°ƒä½¿ç”¨
        self._print_params = {
            "out_path": out,
            "open_only": self.cb_pr.currentIndex() == 0,
            "copies": self.sp_cpy.value(),
            "force_rotate": self.chk_rotate.isChecked()
        }
        
        # ä½¿ç”¨å¼‚æ­¥ PDF åˆå¹¶
        files = [x["p"] for x in self.data]
        self.pdf_worker = PdfWorker(files, m, paper, o, self.chk_cut.isChecked(), out, self)
        self.pdf_worker.progress.connect(self._on_pdf_progress)
        self.pdf_worker.finished.connect(self._on_pdf_merge_finished)
        self.pdf_worker.error.connect(self._on_pdf_error)
        self.pdf_worker.start()
    
    def _on_pdf_progress(self, current, total):
        """PDF åˆå¹¶è¿›åº¦"""
        self.btn_go.setText(f"åˆå¹¶ PDF ({current}/{total})...")
        QApplication.processEvents()
    
    def _on_pdf_merge_finished(self, out_path):
        """PDF åˆå¹¶å®Œæˆï¼Œå¼€å§‹æ‰“å°"""
        self.pdf_worker = None
        params = self._print_params
        
        if params["open_only"]:
            # ä»…æ‰“å¼€ PDF
            if platform.system() == "Windows":
                os.startfile(out_path, "print")
            elif platform.system() == "Darwin":
                os.system(f"open '{out_path}'")
            else:
                os.system(f"xdg-open '{out_path}'")
            self.btn_go.setText(" å¼€å§‹æ‰“å°")
            self.btn_go.setEnabled(True)
        else:
            # å¼‚æ­¥æ‰“å°
            self._start_async_print(out_path, params["copies"], params["force_rotate"])
    
    def _on_pdf_error(self, error_msg):
        """PDF åˆå¹¶é”™è¯¯"""
        self.pdf_worker = None
        self.btn_go.setText("é‡è¯•")
        self.btn_go.setEnabled(True)
        QMessageBox.critical(self, "Error", error_msg)
    
    def _start_async_print(self, pdf_path, copies, force_rotate):
        """å¯åŠ¨å¼‚æ­¥æ‰“å°"""
        p_name = self.cb_pr.currentText().replace("ğŸ–¨ï¸ ", "")
        self.btn_go.setText(f"æ­£åœ¨å‘é€è‡³ {p_name}...")
        
        self.print_worker = PrintWorker(pdf_path, self.current_printer, copies, force_rotate, self)
        self.print_worker.progress.connect(self._on_print_progress)
        self.print_worker.finished.connect(self._on_print_finished)
        self.print_worker.start()
    
    def _on_print_progress(self, current, total):
        """æ‰“å°è¿›åº¦"""
        self.btn_go.setText(f"æ‰“å°ä¸­ ({current}/{total})...")
        QApplication.processEvents()
    
    def _on_print_finished(self, success, msg):
        """æ‰“å°å®Œæˆ"""
        self.print_worker = None
        self.btn_go.setText(" å¼€å§‹æ‰“å°")
        self.btn_go.setEnabled(True)
        
        if success:
            QMessageBox.information(self, "å®Œæˆ", "å·²å‘é€")
        else:
            QMessageBox.critical(self, "é”™è¯¯", msg)
