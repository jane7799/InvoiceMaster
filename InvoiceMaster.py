import sys
import os
import re
import datetime
import base64
import time
import requests
import pandas as pd
import fitz  # PyMuPDF
import platform
import gc
import warnings
import logging
from logging.handlers import RotatingFileHandler
from license_manager import LicenseManager

# 屏蔽 SSL 警告
warnings.filterwarnings("ignore", category=UserWarning, module='urllib3')

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QLabel, QFileDialog, 
                             QListWidget, QListWidgetItem, QMessageBox, QSplitter,
                             QComboBox, QSpinBox, QGroupBox, QSplashScreen, 
                             QDialog, QLineEdit, QFormLayout, QFrame, QCheckBox,
                             QRadioButton, QButtonGroup, QAbstractItemView, 
                             QGraphicsDropShadowEffect, QSizePolicy, QMenu, 
                             QScrollArea, QStackedWidget, QInputDialog, QProgressBar,
                             QToolButton)
from PyQt6.QtCore import Qt, QSettings, QSize, QMimeData, pyqtSignal, QByteArray, QRectF, QPointF, QTimer, QEvent, QPropertyAnimation, QEasingCurve, QPoint
from PyQt6.QtGui import QPixmap, QIcon, QDragEnterEvent, QDropEvent, QImage, QColor, QPainter, QPalette, QPen, QFont, QAction, QCursor, QPageLayout, QTransform
from PyQt6.QtPrintSupport import QPrinterInfo, QPageSetupDialog, QPrinter, QPrintDialog

# ==========================================
# 全局配置
# ==========================================
APP_VERSION = "V1.0.0"
APP_NAME = "智能发票打印助手"
APP_AUTHOR_CN = "© 会钓鱼的猫" 

# ==========================================
# 平台自适应UI配置
# ==========================================
def _detect_platform():
    """检测平台并返回UI配置"""
    system = platform.system()
    
    # 默认配置(现代平台)
    config = {
        "use_animations": True,
        "use_gradients": True,
        "shadow_blur": 25,
        "shadow_opacity": 25,
        "is_legacy": False,
        "platform_name": system
    }
    
    if system == "Windows":
        try:
            import sys
            ver = sys.getwindowsversion()
            # Win7: major=6, minor=1
            # Win10: major=10
            if ver.major == 6 and ver.minor <= 1:
                # Windows 7 或更早版本 - 简化UI
                config.update({
                    "use_animations": False,
                    "use_gradients": False,
                    "shadow_blur": 10,
                    "shadow_opacity": 15,
                    "is_legacy": True,
                    "platform_name": "Windows 7"
                })
            else:
                config["platform_name"] = "Windows 10+"
        except:
            pass
    elif system == "Linux":
        # 统信等Linux系统 - 中等配置
        config.update({
            "use_animations": True,
            "use_gradients": True,
            "shadow_blur": 15,
            "shadow_opacity": 20,
            "platform_name": "Linux/UOS"
        })
    
    return config

# 初始化平台配置
UI_CONFIG = _detect_platform() 

def resource_path(relative_path):
    try: base_path = sys._MEIPASS
    except Exception: base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# ==========================================
# 日志管理器
# ==========================================
class LogManager:
    @staticmethod
    def get_log_directory():
        """获取跨平台日志目录"""
        system = platform.system()
        if system == "Windows":
            base = os.environ.get('APPDATA', os.path.expanduser('~'))
            return os.path.join(base, 'InvoiceMaster', 'logs')
        elif system == "Darwin":  # macOS
            return os.path.expanduser('~/Library/Logs/InvoiceMaster')
        else:  # Linux/UOS
            return os.path.expanduser('~/.local/share/InvoiceMaster/logs')
    
    @staticmethod
    def setup_logging():
        """配置日志系统"""
        try:
            # 创建日志目录
            log_dir = LogManager.get_log_directory()
            os.makedirs(log_dir, exist_ok=True)
            
            log_file = os.path.join(log_dir, 'invoice_master.log')
            
            # 文件处理器（带轮转）
            file_handler = RotatingFileHandler(
                log_file,
                maxBytes=10*1024*1024,  # 10 MB
                backupCount=5,
                encoding='utf-8'
            )
            file_handler.setLevel(logging.INFO)
            
            # 控制台处理器
            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.WARNING)  # 控制台只显示警告及以上
            
            # 设置格式
            formatter = logging.Formatter(
                '%(asctime)s [%(levelname)s] %(name)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            file_handler.setFormatter(formatter)
            console_handler.setFormatter(formatter)
            
            # 配置根日志记录器
            logger = logging.getLogger()
            logger.setLevel(logging.INFO)
            logger.addHandler(file_handler)
            logger.addHandler(console_handler)
            
            return logger
        except Exception as e:
            print(f"日志系统初始化失败: {e}")
            return logging.getLogger()

# ==========================================
# 0. 图标资源
# ==========================================
class Icons:
    _SVGS = {
        "upload": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><path d="M21 15v4a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2v-4"/><polyline points="17 8 12 3 7 8"/><line x1="12" y1="3" x2="12" y2="15"/></svg>""",
        "settings": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><circle cx="12" cy="12" r="3"/><path d="M19.4 15a1.65 1.65 0 0 0 .33 1.82l.06.06a2 2 0 0 1 0 2.83 2 2 0 0 1-2.83 0l-.06-.06a1.65 1.65 0 0 0-1.82-.33 1.65 1.65 0 0 0-1 1.51V21a2 2 0 0 1-2 2 2 2 0 0 1-2-2v-.09A1.65 1.65 0 0 0 9 19.4a1.65 1.65 0 0 0-1.82.33l-.06.06a2 2 0 0 1-2.83 0 2 2 0 0 1 0-2.83l.06-.06a1.65 1.65 0 0 0 .33-1.82 1.65 1.65 0 0 0-1.51-1H3a2 2 0 0 1-2-2 2 2 0 0 1 2-2h.09A1.65 1.65 0 0 0 4.6 9a1.65 1.65 0 0 0-.33-1.82l-.06-.06a2 2 0 0 1 0-2.83 2 2 0 0 1 2.83 0l.06.06a1.65 1.65 0 0 0 1.82.33H9a1.65 1.65 0 0 0 1-1.51V3a2 2 0 0 1 2-2 2 2 0 0 1 2 2v.09a1.65 1.65 0 0 0 1 1.51 1.65 1.65 0 0 0 1.82-.33l.06-.06a2 2 0 0 1 2.83 0 2 2 0 0 1 0 2.83l-.06.06a1.65 1.65 0 0 0-.33 1.82V9a1.65 1.65 0 0 0 1.51 1H21a2 2 0 0 1 2 2 2 2 0 0 1-2 2h-.09a1.65 1.65 0 0 0-1.51 1z"/></svg>""",
        "trash": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><polyline points="3 6 5 6 21 6"/><path d="M19 6v14a2 2 0 0 1-2 2H7a2 2 0 0 1-2-2V6m3 0V4a2 2 0 0 1 2-2h4a2 2 0 0 1 2 2v2"/><line x1="10" y1="11" x2="10" y2="17"/><line x1="14" y1="11" x2="14" y2="17"/></svg>""",
        "excel": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>""",
        "print": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><polyline points="6 9 6 2 18 2 18 9"/><path d="M6 18H4a2 2 0 0 1-2-2v-5a2 2 0 0 1 2-2h16a2 2 0 0 1 2 2v5a2 2 0 0 1-2 2h-2"/><rect x="6" y="14" width="12" height="8"/></svg>""",
        "file": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><path d="M14 2H6a2 2 0 0 0-2 2v16a2 2 0 0 0 2 2h12a2 2 0 0 0 2-2V8z"/><polyline points="14 2 14 8 20 8"/><line x1="16" y1="13" x2="8" y2="13"/><line x1="16" y1="17" x2="8" y2="17"/><polyline points="10 9 9 9 8 9"/></svg>""",
        "prev": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><polyline points="15 18 9 12 15 6"/></svg>""",
        "next": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><polyline points="9 18 15 12 9 6"/></svg>""",
        "zoom_in": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/><line x1="11" y1="8" x2="11" y2="14"/><line x1="8" y1="11" x2="14" y2="11"/></svg>""",
        "zoom_out": """<svg viewBox="0 0 24 24" fill="none" stroke="{c}" stroke-width="2"><circle cx="11" cy="11" r="8"/><line x1="21" y1="21" x2="16.65" y2="16.65"/><line x1="8" y1="11" x2="14" y2="11"/></svg>""",
        "layout_1x1_card": """<svg viewBox="0 0 48 48" fill="none"><rect x="8" y="6" width="32" height="36" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><text x="50%" y="50%" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="10">发票</text></svg>""",
        "layout_1x2_card_v": """<svg viewBox="0 0 48 48" fill="none"><rect x="6" y="6" width="16" height="36" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><rect x="26" y="6" width="16" height="36" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><text x="14" y="24" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="8">发票</text><text x="34" y="24" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="8">发票</text></svg>""",
        "layout_1x2_card_h": """<svg viewBox="0 0 48 48" fill="none"><rect x="6" y="6" width="36" height="16" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><rect x="6" y="26" width="36" height="16" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><text x="24" y="14" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="8">发票</text><text x="24" y="34" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="8">发票</text></svg>""",
        "layout_2x2_card": """<svg viewBox="0 0 48 48" fill="none"><rect x="6" y="6" width="16" height="16" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><rect x="26" y="6" width="16" height="16" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><rect x="6" y="26" width="16" height="16" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><rect x="26" y="26" width="16" height="16" rx="2" fill="#f0f0f0" stroke="{c}" stroke-width="2"/><text x="14" y="14" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="6">发票</text><text x="34" y="14" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="6">发票</text><text x="14" y="34" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="6">发票</text><text x="34" y="34" dominant-baseline="middle" text-anchor="middle" fill="#999" font-size="6">发票</text></svg>"""
    }
    @staticmethod
    def get(name, color="#555"):
        return QIcon(QPixmap.fromImage(QImage.fromData(QByteArray(Icons._SVGS.get(name,"").format(c=color).encode()))))

# ==========================================
# 1. 现代化主题系统
# ==========================================
class ThemeManager:
    SCROLLBAR_CSS = """
        QScrollBar:vertical {
            border: none;
            background: transparent;
            width: 10px;
            margin: 0px;
        }
        QScrollBar::handle:vertical {
            background: rgba(0, 0, 0, 0.2);
            min-height: 30px;
            border-radius: 5px;
            margin: 2px;
        }
        QScrollBar::handle:vertical:hover {
            background: rgba(0, 0, 0, 0.3);
        }
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
            height: 0px;
        }
        QScrollBar:horizontal {
            border: none;
            background: transparent;
            height: 10px;
            margin: 0px;
        }
        QScrollBar::handle:horizontal {
            background: rgba(0, 0, 0, 0.2);
            min-width: 30px;
            border-radius: 5px;
            margin: 2px;
        }
        QScrollBar::handle:horizontal:hover {
            background: rgba(0, 0, 0, 0.3);
        }
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
            width: 0px;
        }
    """
    
    COMMON_CSS = """
        QRadioButton {
            color: #475569;
            font-weight: 500;
            font-size: 13px;
            spacing: 8px;
        }
        QCheckBox {
            color: #475569;
            font-weight: 500;
            font-size: 13px;
            spacing: 8px;
        }
        QWidget#ItemRow {
            background: transparent;
        }
        QLabel#ItemTitle {
            font-weight: 600;
            font-size: 13px;
            color: #1E293B;
        }
        QLabel#ItemDetail {
            color: #64748B;
            font-size: 12px;
        }
        QPushButton#RowDelBtn {
            background: transparent;
            border: none;
            border-radius: 4px;
        }
        QPushButton#RowDelBtn:hover {
            background-color: #FEE2E2;
        }
        
        QToolButton#LayoutCard {
            background-color: white;
            border: 2px solid #CBD5E1;
            border-radius: 12px;
            padding: 4px;
        }
        QToolButton#LayoutCard:hover {
            border-color: #3B82F6;
            background-color: #EFF6FF;
            border-width: 2px;
        }
        QToolButton#LayoutCard:checked {
            border: 3px solid #2563EB;
            background-color: #DBEAFE;
        }
        QFrame#PreviewControlBar {
            background-color: white;
            border-top: 1px solid #E2E8F0;
        }
        QLabel#PageLabel {
            font-size: 12px;
            color: #64748B;
        }
        QLabel#Title {
            font-size: 15px;
            font-weight: 600;
            color: #1E293B;
        }
    """

    CSS_LIGHT = """
    QMainWindow {
        background-color: #F8FAFC;
    }
    QWidget {
        color: #1E293B;
        font-family: "PingFang SC", "Microsoft YaHei UI", "Segoe UI", sans-serif;
        font-size: 13px;
    }
    QFrame#Card {
        background-color: white;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
    }
    QPushButton {
        background-color: white;
        border: 1.5px solid #CBD5E1;
        border-radius: 6px;
        padding: 8px 16px;
        font-weight: 500;
        color: #475569;
    }
    QPushButton:hover {
        background-color: #F8FAFC;
        border-color: #2563EB;
        color: #2563EB;
    }
    QPushButton:pressed {
        background-color: #EFF6FF;
    }
    QPushButton#IconBtn {
        background: transparent;
        border: none;
        border-radius: 6px;
        padding: 6px;
    }
    QPushButton#IconBtn:hover {
        background-color: #F1F5F9;
    }
    QPushButton#PrimaryBtn {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 #3B82F6, stop:1 #2563EB);
        border: none;
        color: white;
        font-weight: 600;
        font-size: 14px;
        border-radius: 8px;
        padding: 10px 20px;
    }
    QPushButton#PrimaryBtn:hover {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 #2563EB, stop:1 #1D4ED8);
    }
    QPushButton#DangerBtn {
        color: #DC2626;
    }
    QPushButton#PropBtn {
        color: #475569;
        border: 1.5px solid #CBD5E1;
    }
    QListWidget {
        background-color: white;
        border: 1px solid #E2E8F0;
        border-radius: 8px;
        outline: none;
        padding: 4px;
    }
    QListWidget::item {
        border-bottom: 1px solid #F1F5F9;
        border-radius: 4px;
        margin: 2px 0px;
    }
    QListWidget::item:selected {
        background-color: #EFF6FF;
        border-left: 3px solid #2563EB;
        color: #1E293B;
    }
    QListWidget::item:hover {
        background-color: #F8FAFC;
    }
    QLineEdit, QComboBox, QSpinBox {
        border: 1.5px solid #CBD5E1;
        border-radius: 6px;
        padding: 8px 12px;
        background: white;
        min-height: 20px;
        font-size: 13px;
    }
    QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
        border-color: #2563EB;
        background: white;
    }
    QGroupBox {
        border: none;
        font-weight: 600;
        margin-top: 12px;
        color: #1E293B;
    }
    """ + SCROLLBAR_CSS + COMMON_CSS
    
    CSS_DARK = """
    QMainWindow {
        background-color: #0F172A;
    }
    QWidget {
        color: #E2E8F0;
        font-family: "Microsoft YaHei UI", sans-serif;
        font-size: 13px;
    }
    QFrame#Card {
        background-color: #1E293B;
        border: 1px solid #334155;
        border-radius: 12px;
    }
    QPushButton {
        background-color: #1E293B;
        border: 1.5px solid #475569;
        border-radius: 6px;
        padding: 8px 16px;
        color: #E2E8F0;
        font-weight: 500;
    }
    QPushButton:hover {
        background-color: #334155;
        border-color: #3B82F6;
        color: white;
    }
    QPushButton#IconBtn {
        background: transparent;
        border: none;
        border-radius: 6px;
    }
    QPushButton#IconBtn:hover {
        background-color: #334155;
    }
    QPushButton#PrimaryBtn {
        background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
            stop:0 #3B82F6, stop:1 #2563EB);
        border: none;
        color: white;
        font-weight: 600;
    }
    QPushButton#DangerBtn {
        color: #EF4444;
    }
    QListWidget {
        background-color: #1E293B;
        border: 1px solid #334155;
        border-radius: 8px;
        outline: none;
        padding: 4px;
    }
    QListWidget::item {
        border-bottom: 1px solid #334155;
        border-radius: 4px;
        margin: 2px 0px;
    }
    QListWidget::item:selected {
        background-color: #1E40AF;
        border-left: 3px solid #3B82F6;
        color: white;
    }
    QListWidget::item:hover {
        background-color: #334155;
    }
    QLineEdit, QComboBox, QSpinBox {
        border: 1.5px solid #475569;
        border-radius: 6px;
        padding: 8px 12px;
        background: #1E293B;
        color: white;
    }
    QLineEdit:focus, QComboBox:focus, QSpinBox:focus {
        border-color: #3B82F6;
    }
    QRadioButton {
        color: #E2E8F0;
    }
    QCheckBox {
        color: #E2E8F0;
    }
    QLabel#ItemTitle {
        color: #F1F5F9;
    }
    QLabel#ItemDetail {
        color: #94A3B8;
    }
    QToolButton#LayoutCard {
        background-color: #1E293B;
        border: 2px solid #475569;
    }
    QToolButton#LayoutCard:checked {
        border-color: #3B82F6;
        background-color: #1E40AF;
    }
    """ + SCROLLBAR_CSS

    @staticmethod
    def apply(app, mode="Light"):
        if mode == "Auto":
            mode = "Light"
        if mode == "Dark":
            app.setStyleSheet(ThemeManager.CSS_DARK)
            return "#ffffff"
        else:
            app.setStyleSheet(ThemeManager.CSS_LIGHT)
            return "#555555"

# ==========================================
# 2. 核心 Helper/Engine 类
# ==========================================
class InvoiceHelper:
    @staticmethod
    def thumb(fp):
        try: 
            with fitz.open(fp) as doc:
                return QPixmap.fromImage(QImage.fromData(doc.load_page(0).get_pixmap(matrix=fitz.Matrix(0.3,0.3)).tobytes("ppm")))
        except: return Icons.get("file", "#ccc").pixmap(100,100)
    @staticmethod
    def parse_amount_local(file_path):
        amount = 0.0; date = ""
        try:
            with fitz.open(file_path) as doc:
                text = "".join([page.get_text() for page in doc])
                m_total = re.search(r'(小写).*?[¥￥]?\s*([0-9,]+\.\d{2})', text)
                if m_total: amount = float(m_total.group(2).replace(",", ""))
                else:
                    amounts = re.findall(r'[¥￥]\s*([0-9,]+\.\d{2})', text); 
                    if amounts: amount = max([float(x.replace(",", "")) for x in amounts])
                m_date = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', text)
                if m_date: date = f"{m_date.group(1)}-{m_date.group(2).zfill(2)}-{m_date.group(3).zfill(2)}"
        except: pass
        return amount, date
    @staticmethod
    def ocr(fp, ak, sk):
        logger = logging.getLogger(__name__)
        if not ak:
            logger.warning("OCR 未配置 API Key")
            return {}
        try:
            logger.info(f"OCR 识别开始: {os.path.basename(fp)}")
            t = requests.get(f"https://aip.baidubce.com/oauth/2.0/token?grant_type=client_credentials&client_id={ak}&client_secret={sk}").json().get("access_token")
            with open(fp,'rb') as f: b = base64.b64encode(f.read()).decode()
            r = requests.post(f"https://aip.baidubce.com/rest/2.0/ocr/v1/vat_invoice?access_token={t}", data={"image":b}, headers={'content-type':'application/x-www-form-urlencoded'}).json()
            wr = r.get("words_result", {})
            items = wr.get("CommodityName", [])
            item_str = ",".join([x.get("word","") for x in items]) if isinstance(items, list) else str(items)
            # 获取税率列表
            tax_rates = wr.get("CommodityTaxRate", [])
            tax_rate_str = ",".join([x.get("word","") for x in tax_rates]) if isinstance(tax_rates, list) else str(tax_rates)
            # 扩展字段:18个专业字段
            result = { 
                "date": wr.get("InvoiceDate", ""),  # 开票日期
                "amount": float(wr.get("AmountInFiguers", "0") or "0"),  # 价税合计
                "amount_without_tax": wr.get("TotalAmount", ""),  # 不含税金额
                "tax_amt": wr.get("TotalTax", ""),  # 税额
                "tax_rate": tax_rate_str,  # 税率
                "seller": wr.get("SellerName", ""),  # 销售方名称
                "seller_tax_id": wr.get("SellerRegisterNum", ""),  # 销售方税号
                "buyer": wr.get("PurchaserName", ""),  # 购买方名称
                "buyer_tax_id": wr.get("PurchaserRegisterNum", ""),  # 购买方税号
                "code": wr.get("InvoiceCode", ""),  # 发票代码
                "number": wr.get("InvoiceNum", ""),  # 发票号码
                "check_code": wr.get("CheckCode", ""),  # 校验码
                "invoice_type": wr.get("InvoiceType", ""),  # 发票类型
                "item_name": item_str,  # 商品明细
                "remark": wr.get("Remarks", ""),  # 备注
                "machine_code": wr.get("MachineCode", ""),  # 机器编号
            }
            logger.info(f"OCR 识别成功: {os.path.basename(fp)}, 金额: {result.get('amount', 0)}")
            return result
        except Exception as e:
            logger.error(f"OCR 识别失败: {os.path.basename(fp)}, 错误: {str(e)}")
            return {}

class PDFEngine:
    SIZES = {"A4":(595,842), "A5":(420,595), "B5":(499,709)}
    @staticmethod
    def merge(files, mode="1x1", paper="A4", orient="V", cutline=True, out_path=None):
        doc = fitz.open(); bw, bh = PDFEngine.SIZES.get(paper, (595,842))
        
        # [V3.3.0] 永远纵向
        PW, PH = (bw, bh) 
        
        cells = []
        if mode == "1x1":
            cells = [(0, 0, PW, PH)] 
        elif mode == "1x2":
            cells = [(0, 0, PW, PH/2), (0, PH/2, PW, PH/2)]
        elif mode == "2x2":
            mw, mh = PW/2, PH/2
            cells = [(0, 0, mw, mh), (mw, 0, mw, mh), (0, mh, mw, mh), (mw, mh, mw, mh)]
            
        if not files: doc.new_page(width=PW, height=PH)
        PADDING = 40
        chunk_size = len(cells)
        try:
            for i in range(0, len(files), chunk_size):
                chunk = files[i:i+chunk_size]; pg = doc.new_page(width=PW, height=PH)
                for j, f in enumerate(chunk):
                    if j >= len(cells): break
                    try:
                        cx, cy, cw, ch = cells[j]
                        rotate_angle = -90 if orient == "H" else 0
                        target_rect = fitz.Rect(cx + PADDING, cy + PADDING, cx + cw - PADDING, cy + ch - PADDING)
                        
                        # 检查文件类型
                        if f.lower().endswith(('.jpg', '.jpeg', '.png')):
                            # 图片文件:直接插入图片
                            pg.insert_image(target_rect, filename=f, keep_proportion=True, rotate=rotate_angle)
                        else:
                            # PDF文件:使用 show_pdf_page
                            with fitz.open(f) as src_doc:
                                pg.show_pdf_page(target_rect, src_doc, 0, keep_proportion=True, rotate=rotate_angle)
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"处理文件失败 {os.path.basename(f)}: {str(e)}")
                        pass
                if cutline:
                    s = pg.new_shape(); s.draw_rect(fitz.Rect(0,0,0,0)) 
                    if mode == "1x2":
                        s.draw_line(fitz.Point(0, PH/2), fitz.Point(PW, PH/2)) 
                    elif mode == "2x2":
                        s.draw_line(fitz.Point(PW/2, 0), fitz.Point(PW/2, PH)) 
                        s.draw_line(fitz.Point(0, PH/2), fitz.Point(PW, PH/2)) 
                    s.finish(color=(0,0,0), width=0.5, dashes=[4,4], stroke_opacity=0.6); s.commit(overlay=True)
            if out_path: doc.save(out_path)
            return doc if not out_path else None
        finally:
            if out_path: doc.close()

class PrinterEngine:
    @staticmethod
    def print_pdf(pdf_path, printer, copies=1, force_rotate=False):
        logger = logging.getLogger(__name__)
        logger.info(f"开始打印: {os.path.basename(pdf_path)}, 份数: {copies}, DPI: 600")
        try:
            # 【V5.0】提高打印质量: DPI 600, 渲染4.0倍
            printer.setCopyCount(copies); printer.setResolution(600); printer.setFullPage(True)
            with fitz.open(pdf_path) as doc:
                painter = QPainter()
                if not painter.begin(printer):
                    logger.error("无法启动打印任务")
                    return False, "无法启动打印任务"
                for i, page in enumerate(doc):
                    if i > 0: printer.newPage()
                    # 【V5.0】提高渲染质量: 4.0倍缩放,保持PDF原始高分辨率
                    pix = page.get_pixmap(matrix=fitz.Matrix(4.0, 4.0), alpha=False)
                    img = QImage.fromData(pix.tobytes("ppm"))
                    
                    # 永远纵向
                    printer.setPageOrientation(QPageLayout.Orientation.Portrait)
                    
                    if force_rotate:
                         transform = QTransform()
                         transform.rotate(90)
                         img = img.transformed(transform)
                    
                    page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
                    
                    # [V3.5.0] 移除 0.96 缩放，全尺寸绘制
                    try: safe_rect = printer.pageLayout().paintRectPixels(printer.resolution())
                    except: safe_rect = page_rect

                    target_w = int(safe_rect.width()); target_h = int(safe_rect.height())
                    # 【V5.0】使用 SmoothTransformation 获得最佳质量
                    scaled_img = img.scaled(target_w, target_h, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
                    
                    x = int(safe_rect.x() + (safe_rect.width() - scaled_img.width()) / 2)
                    y = int(safe_rect.y() + (safe_rect.height() - scaled_img.height()) / 2)
                    
                    painter.drawImage(x, y, scaled_img)
                painter.end()
            logger.info(f"打印任务发送成功: {os.path.basename(pdf_path)}")
            return True, "发送成功"
        except Exception as e:
            logger.error(f"打印失败: {os.path.basename(pdf_path)}, 错误: {str(e)}", exc_info=True)
            return False, str(e)

# ==========================================
# 3. UI 组件
# ==========================================
class Card(QFrame):
    def __init__(self):
        super().__init__(); self.setObjectName("Card")
        # 使用平台自适应阴影配置
        blur = UI_CONFIG.get("shadow_blur", 25)
        opacity = UI_CONFIG.get("shadow_opacity", 25)
        eff = QGraphicsDropShadowEffect(); eff.setBlurRadius(blur); eff.setColor(QColor(0,0,0,opacity)); eff.setOffset(0,4 if not UI_CONFIG.get("is_legacy") else 2); self.setGraphicsEffect(eff)

class DragArea(QLabel):
    dropped = pyqtSignal(list)
    def __init__(self):
        super().__init__()
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setAcceptDrops(True)
        self.setMinimumHeight(120)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self._hover = False
        self._update_style()
        
    def _update_style(self):
        use_gradients = UI_CONFIG.get("use_gradients", True)
        
        if self._hover:
            if use_gradients:
                bg = """background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 rgba(37, 99, 235, 0.15),
                        stop:1 rgba(37, 99, 235, 0.05));"""
            else:
                bg = "background: rgba(37, 99, 235, 0.1);"
            self.setStyleSheet(f"""
                QLabel {{
                    border: 2px dashed #2563EB;
                    border-radius: 12px;
                    {bg}
                }}
            """)
        else:
            if use_gradients:
                bg = """background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 rgba(148, 163, 184, 0.08),
                        stop:1 rgba(148, 163, 184, 0.02));"""
            else:
                bg = "background: rgba(148, 163, 184, 0.05);"
            self.setStyleSheet(f"""
                QLabel {{
                    border: 2px dashed #94A3B8;
                    border-radius: 12px;
                    {bg}
                }}
            """)
    
    def upd(self, c): 
        self.setPixmap(Icons.get("upload", c if not self._hover else "#2563EB").pixmap(56, 56))
    
    def enterEvent(self, e):
        self._hover = True
        self._update_style()
        self.upd("#2563EB")
        
    def leaveEvent(self, e):
        self._hover = False
        self._update_style()
        self.upd("#94A3B8")
        
    def dragEnterEvent(self, e):
        self._hover = True
        self._update_style()
        e.accept()
        
    def dragLeaveEvent(self, e):
        self._hover = False
        self._update_style()
        
    def dropEvent(self, e):
        self._hover = False
        self._update_style()
        self.dropped.emit([u.toLocalFile() for u in e.mimeData().urls() if u.toLocalFile().lower().endswith(('.pdf','.jpg','.png'))])
        
    def mousePressEvent(self, e): 
        fs, _ = QFileDialog.getOpenFileNames(self, "添加发票", "", "发票文件 (*.pdf *.jpg *.png)")
        if fs: self.dropped.emit(fs)

class InvoiceItemWidget(QWidget):
    def __init__(self, data, parent_item, delete_callback):
        super().__init__()
        self.data = data
        self.parent_item = parent_item
        self.delete_callback = delete_callback
        self.setObjectName("ItemRow")
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 8, 5, 8)
        layout.setSpacing(10)
        
        self.icon_lbl = QLabel()
        self.icon_lbl.setPixmap(Icons.get("file", "#888").pixmap(32, 32))
        layout.addWidget(self.icon_lbl)
        
        text_layout = QVBoxLayout()
        text_layout.setSpacing(2)
        
        # 标题行（包含文件名和状态标识）
        title_row = QHBoxLayout()
        title_row.setSpacing(6)
        
        self.lbl_title = QLabel(data['n'])
        self.lbl_title.setObjectName("ItemTitle")
        title_row.addWidget(self.lbl_title)
        
        # 状态标识
        self.status_badge = QLabel()
        self.status_badge.setFixedSize(18, 18)
        self.status_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_badge.hide()
        title_row.addWidget(self.status_badge)
        title_row.addStretch()
        
        text_layout.addLayout(title_row)
        
        self.lbl_detail = QLabel(f"{data['d']} | ¥{data['a']:.2f}")
        self.lbl_detail.setObjectName("ItemDetail")
        text_layout.addWidget(self.lbl_detail)
        
        layout.addLayout(text_layout)
        layout.addStretch()
        
        self.btn_del = QPushButton()
        self.btn_del.setObjectName("RowDelBtn")
        self.btn_del.setIcon(Icons.get("trash", "#d73a49"))
        self.btn_del.setFixedSize(28, 28)
        self.btn_del.setToolTip("删除此发票")
        self.btn_del.clicked.connect(self.on_delete_clicked)
        layout.addWidget(self.btn_del)
        
        self.update_status_badge()
    
    def on_delete_clicked(self):
        self.delete_callback(self.parent_item)
    
    def update_display(self, new_data):
        self.data = new_data
        self.lbl_title.setText(new_data['n'])
        self.lbl_detail.setText(f"{new_data['d']} | ¥{new_data['a']:.2f}")
        self.update_status_badge()
    
    def update_status_badge(self):
        """更新状态标识"""
        has_amount = self.data.get('a', 0) > 0
        is_manually_edited = self.data.get('manually_edited', False)
        
        if not has_amount:
            self.status_badge.setText("⚠️")
            self.status_badge.setStyleSheet("background: #FEE2E2; border-radius: 9px; font-size: 12px;")
            self.status_badge.setToolTip("未识别到金额，请手动修改")
            self.status_badge.show()
            self.setStyleSheet("")
        elif is_manually_edited:
            self.status_badge.setText("✓")
            self.status_badge.setStyleSheet("background: #D1FAE5; border-radius: 9px; font-size: 12px; color: #059669; font-weight: bold;")
            self.status_badge.setToolTip("已手动修改金额")
            self.status_badge.show()
            self.setStyleSheet("QWidget#ItemRow { background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 rgba(209, 250, 229, 0.3), stop:1 rgba(209, 250, 229, 0.1)); border-left: 3px solid #10B981; }")
        else:
            self.status_badge.hide()
            self.setStyleSheet("")

class SettingsDlg(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("设置")
        self.resize(550, 420)
        self.setStyleSheet("""
            QDialog {
                background-color: #F8FAFC;
            }
        """)
        self.parent = parent
        s = QSettings("MySoft", "InvoiceMaster")
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # 渐变标题栏
        header = QWidget()
        header.setFixedHeight(90)
        header.setStyleSheet("""
            background: qlineargradient(
                x1:0, y1:0, x2:1, y2:0,
                stop:0 #2563EB,
                stop:1 #7C3AED
            );
        """)
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(30, 20, 30, 20)
        header_layout.setSpacing(5)
        
        title = QLabel("⚙️ 应用设置")
        title.setStyleSheet("""
            font-size: 22px;
            font-weight: 700;
            color: white;
            background: transparent;
        """)
        header_layout.addWidget(title)
        
        subtitle = QLabel("配置主题、API密钥和授权信息")
        subtitle.setStyleSheet("""
            font-size: 13px;
            color: rgba(255, 255, 255, 0.9);
            background: transparent;
        """)
        header_layout.addWidget(subtitle)
        
        layout.addWidget(header)
        
        # 内容区域
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(30, 30, 30, 30)
        content_layout.setSpacing(20)
        
        # 激活状态卡片
        if hasattr(parent, 'license_manager'):
            activation_card = QFrame()
            activation_card.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border-radius: 12px;
                }
            """)
            
            activation_shadow = QGraphicsDropShadowEffect()
            activation_shadow.setBlurRadius(20)
            activation_shadow.setColor(QColor(0, 0, 0, 30))
            activation_shadow.setOffset(0, 2)
            activation_card.setGraphicsEffect(activation_shadow)
            
            activation_layout = QVBoxLayout(activation_card)
            activation_layout.setContentsMargins(20, 20, 20, 20)
            activation_layout.setSpacing(15)
            
            card_title = QLabel("🔑 Excel 导出授权")
            card_title.setStyleSheet("""
                font-size: 15px;
                font-weight: 600;
                color: #1E293B;
            """)
            activation_layout.addWidget(card_title)
            
            info = parent.license_manager.get_activation_info()
            status_container = QHBoxLayout()
            
            if info['is_activated']:
                activation_status = QLabel("✅ 已激活")
                activation_status.setStyleSheet("""
                    color: #10B981;
                    font-weight: 600;
                    font-size: 14px;
                    padding: 8px 16px;
                    background: #ECFDF5;
                    border-radius: 6px;
                """)
            else:
                remaining = info['remaining_trials']
                activation_status = QLabel(f"⚠️ 未激活 (剩余: {remaining}/10次)")
                activation_status.setStyleSheet("""
                    color: #F59E0B;
                    font-weight: 600;
                    font-size: 14px;
                    padding: 8px 16px;
                    background: #FEF3C7;
                    border-radius: 6px;
                """)
            
            activation_btn = QPushButton("激活管理")
            activation_btn.setStyleSheet("""
                QPushButton {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #3B82F6, stop:1 #2563EB);
                    border: none;
                    color: white;
                    font-weight: 600;
                    font-size: 13px;
                    border-radius: 6px;
                    padding: 8px 20px;
                }
                QPushButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #2563EB, stop:1 #1D4ED8);
                }
            """)
            activation_btn.clicked.connect(self.open_activation_dialog)
            
            status_container.addWidget(activation_status)
            status_container.addStretch()
            status_container.addWidget(activation_btn)
            
            activation_layout.addLayout(status_container)
            content_layout.addWidget(activation_card)
        
        # 主题设置卡片
        theme_card = QFrame()
        theme_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        theme_shadow = QGraphicsDropShadowEffect()
        theme_shadow.setBlurRadius(20)
        theme_shadow.setColor(QColor(0, 0, 0, 30))
        theme_shadow.setOffset(0, 2)
        theme_card.setGraphicsEffect(theme_shadow)
        
        theme_layout = QVBoxLayout(theme_card)
        theme_layout.setContentsMargins(20, 20, 20, 20)
        theme_layout.setSpacing(15)
        
        theme_title = QLabel("🎨 外观主题")
        theme_title.setStyleSheet("""
            font-size: 15px;
            font-weight: 600;
            color: #1E293B;
        """)
        theme_layout.addWidget(theme_title)
        
        self.cb_th = QComboBox()
        self.cb_th.addItems(["Light", "Dark"])
        self.cb_th.setCurrentText(s.value("theme", "Light"))
        self.cb_th.setStyleSheet("""
            QComboBox {
                padding: 10px 12px;
                border: 2px solid #CBD5E1;
                border-radius: 8px;
                background: white;
                font-size: 14px;
            }
            QComboBox:focus {
                border-color: #2563EB;
            }
        """)
        self.cb_th.currentTextChanged.connect(parent.change_theme)
        theme_layout.addWidget(self.cb_th)
        
        content_layout.addWidget(theme_card)
        
        # API配置卡片
        api_card = QFrame()
        api_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        api_shadow = QGraphicsDropShadowEffect()
        api_shadow.setBlurRadius(20)
        api_shadow.setColor(QColor(0, 0, 0, 30))
        api_shadow.setOffset(0, 2)
        api_card.setGraphicsEffect(api_shadow)
        
        api_layout = QVBoxLayout(api_card)
        api_layout.setContentsMargins(20, 20, 20, 20)
        api_layout.setSpacing(15)
        
        api_title = QLabel("🔐 百度 OCR API 配置")
        api_title.setStyleSheet("""
            font-size: 15px;
            font-weight: 600;
            color: #1E293B;
        """)
        api_layout.addWidget(api_title)
        
        link = QLabel('<a href="https://console.bce.baidu.com">点击申请百度OCR Key (免费)</a>')
        link.setOpenExternalLinks(True)
        link.setStyleSheet("color: #2563EB; font-size: 12px;")
        api_layout.addWidget(link)
        
        self.ak = QLineEdit(s.value("ak", ""))
        self.ak.setPlaceholderText("API Key")
        self.ak.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 2px solid #CBD5E1;
                border-radius: 8px;
                background: white;
                font-size: 14px;
            }
            QLineEdit:focus {
                border-color: #2563EB;
            }
        """)
        api_layout.addWidget(self.ak)
        
        self.sk = QLineEdit(s.value("sk", ""))
        self.sk.setEchoMode(QLineEdit.EchoMode.Password)
        self.sk.setPlaceholderText("Secret Key")
        self.sk.setStyleSheet("""
            QLineEdit {
                padding: 10px 12px;
                border: 2px solid #CBD5E1;
                border-radius: 8px;
                background: white;
                font-size: 14px;
            }
            QLineEdit:focus {
                border-color: #2563EB;
            }
        """)
        api_layout.addWidget(self.sk)
        
        content_layout.addWidget(api_card)
        content_layout.addStretch()
        
        layout.addWidget(content)
        
        # 底部按钮区域
        button_container = QWidget()
        button_container.setStyleSheet("background-color: white; border-top: 1px solid #E2E8F0;")
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(30, 20, 30, 20)
        button_layout.setSpacing(12)
        
        button_layout.addStretch()
        
        save_btn = QPushButton("💾 保存配置")
        save_btn.setMinimumHeight(44)
        save_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3B82F6, stop:1 #2563EB);
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 32px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2563EB, stop:1 #1D4ED8);
            }
        """)
        save_btn.clicked.connect(self.save)
        button_layout.addWidget(save_btn)
        
        layout.addWidget(button_container)
    
    def open_activation_dialog(self):
        """打开激活管理对话框"""
        dialog = ActivationDialog(self, self.parent.license_manager)
        dialog.exec()
    
    def save(self):
        s = QSettings("MySoft", "InvoiceMaster")
        s.setValue("ak", self.ak.text())
        s.setValue("sk", self.sk.text())
        s.setValue("theme", self.cb_th.currentText())
        self.accept()


class ActivationDialog(QDialog):
    """激活管理对话框"""
    def __init__(self, parent, license_manager):
        super().__init__(parent)
        self.license_manager = license_manager
        self.setWindowTitle("激活管理")
        self.resize(550, 520)
        self.setStyleSheet("""
            QDialog {
                background-color: #F8FAFC;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # 渐变标题栏
        header = QWidget()
        header.setFixedHeight(100)
        header.setStyleSheet("""
            background: qlineargradient(
                x1:0, y1:0, x2:1, y2:0,
                stop:0 #2563EB,
                stop:1 #7C3AED
            );
            border-radius: 0px;
        """)
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(30, 20, 30, 20)
        header_layout.setSpacing(5)
        
        title = QLabel("🔑 Excel 导出功能激活")
        title.setStyleSheet("""
            font-size: 22px;
            font-weight: 700;
            color: white;
            background: transparent;
        """)
        header_layout.addWidget(title)
        
        subtitle = QLabel("管理您的软件授权和试用状态")
        subtitle.setStyleSheet("""
            font-size: 13px;
            color: rgba(255, 255, 255, 0.9);
            background: transparent;
        """)
        header_layout.addWidget(subtitle)
        
        layout.addWidget(header)
        
        # 内容区域
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(30, 30, 30, 30)
        content_layout.setSpacing(20)
        
        # 激活状态卡片
        info = self.license_manager.get_activation_info()
        status_card = QFrame()
        status_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        status_shadow = QGraphicsDropShadowEffect()
        status_shadow.setBlurRadius(20)
        status_shadow.setColor(QColor(0, 0, 0, 30))
        status_shadow.setOffset(0, 2)
        status_card.setGraphicsEffect(status_shadow)
        
        status_layout = QVBoxLayout(status_card)
        status_layout.setContentsMargins(20, 20, 20, 20)
        status_layout.setSpacing(10)
        
        if info['is_activated']:
            status_text = "✅ 已激活"
            status_color = "#10B981"
            status_bg = "#ECFDF5"
        else:
            remaining = info['remaining_trials']
            status_text = f"⚠️ 未激活 (剩余试用: {remaining}/10次)"
            status_color = "#F59E0B"
            status_bg = "#FEF3C7"
        
        self.status_label = QLabel(status_text)
        self.status_label.setStyleSheet(f"""
            font-size: 16px;
            font-weight: 600;
            color: {status_color};
            padding: 12px 20px;
            background: {status_bg};
            border-radius: 8px;
        """)
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        status_layout.addWidget(self.status_label)
        
        content_layout.addWidget(status_card)
        
        # 机器码卡片
        machine_card = QFrame()
        machine_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        machine_shadow = QGraphicsDropShadowEffect()
        machine_shadow.setBlurRadius(20)
        machine_shadow.setColor(QColor(0, 0, 0, 30))
        machine_shadow.setOffset(0, 2)
        machine_card.setGraphicsEffect(machine_shadow)
        
        machine_layout = QVBoxLayout(machine_card)
        machine_layout.setContentsMargins(20, 20, 20, 20)
        machine_layout.setSpacing(12)
        
        machine_title = QLabel("📱 机器码")
        machine_title.setStyleSheet("""
            font-size: 15px;
            font-weight: 600;
            color: #1E293B;
        """)
        machine_layout.addWidget(machine_title)
        
        hint = QLabel("请将以下机器码发送给开发者以获取激活码")
        hint.setStyleSheet("""
            color: #64748B;
            font-size: 12px;
        """)
        machine_layout.addWidget(hint)
        
        machine_code_layout = QHBoxLayout()
        self.machine_code_edit = QLineEdit(self.license_manager.get_machine_code())
        self.machine_code_edit.setReadOnly(True)
        self.machine_code_edit.setStyleSheet("""
            font-family: 'Courier New', monospace;
            font-size: 15px;
            font-weight: 600;
            padding: 12px 16px;
            background: #F1F5F9;
            border: 2px solid #E2E8F0;
            border-radius: 8px;
            color: #1E293B;
        """)
        machine_code_layout.addWidget(self.machine_code_edit)
        
        copy_btn = QPushButton("📋 复制")
        copy_btn.setFixedWidth(90)
        copy_btn.setStyleSheet("""
            QPushButton {
                background-color: white;
                border: 1.5px solid #CBD5E1;
                border-radius: 8px;
                padding: 12px 16px;
                font-weight: 500;
                color: #475569;
            }
            QPushButton:hover {
                background-color: #F8FAFC;
                border-color: #2563EB;
                color: #2563EB;
            }
        """)
        copy_btn.clicked.connect(self.copy_machine_code)
        machine_code_layout.addWidget(copy_btn)
        
        machine_layout.addLayout(machine_code_layout)
        content_layout.addWidget(machine_card)
        
        # 激活码卡片
        activation_card = QFrame()
        activation_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        activation_shadow = QGraphicsDropShadowEffect()
        activation_shadow.setBlurRadius(20)
        activation_shadow.setColor(QColor(0, 0, 0, 30))
        activation_shadow.setOffset(0, 2)
        activation_card.setGraphicsEffect(activation_shadow)
        
        activation_layout = QVBoxLayout(activation_card)
        activation_layout.setContentsMargins(20, 20, 20, 20)
        activation_layout.setSpacing(12)
        
        activation_title = QLabel("🔐 激活码")
        activation_title.setStyleSheet("""
            font-size: 15px;
            font-weight: 600;
            color: #1E293B;
        """)
        activation_layout.addWidget(activation_title)
        
        activation_hint = QLabel("请输入从开发者处获得的激活码")
        activation_hint.setStyleSheet("""
            color: #64748B;
            font-size: 12px;
        """)
        activation_layout.addWidget(activation_hint)
        
        self.activation_code_edit = QLineEdit()
        self.activation_code_edit.setPlaceholderText("XXXX-XXXX-XXXX-XXXX")
        self.activation_code_edit.setStyleSheet("""
            font-family: 'Courier New', monospace;
            font-size: 15px;
            font-weight: 600;
            padding: 12px 16px;
            background: white;
            border: 2px solid #CBD5E1;
            border-radius: 8px;
            color: #1E293B;
        """)
        activation_layout.addWidget(self.activation_code_edit)
        
        content_layout.addWidget(activation_card)
        content_layout.addStretch()
        
        layout.addWidget(content)
        
        # 底部按钮区域
        button_container = QWidget()
        button_container.setStyleSheet("background-color: white; border-top: 1px solid #E2E8F0;")
        button_layout = QHBoxLayout(button_container)
        button_layout.setContentsMargins(30, 20, 30, 20)
        button_layout.setSpacing(12)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setMinimumHeight(44)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: white;
                border: 1.5px solid #CBD5E1;
                border-radius: 8px;
                padding: 10px 32px;
                font-weight: 500;
                color: #475569;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #F8FAFC;
                border-color: #2563EB;
                color: #2563EB;
            }
        """)
        cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(cancel_btn)
        
        button_layout.addStretch()
        
        activate_btn = QPushButton("🔓 立即激活")
        activate_btn.setMinimumHeight(44)
        activate_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3B82F6, stop:1 #2563EB);
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 32px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #2563EB, stop:1 #1D4ED8);
            }
        """)
        activate_btn.clicked.connect(self.activate)
        button_layout.addWidget(activate_btn)
        
        layout.addWidget(button_container)
    
    def copy_machine_code(self):
        """复制机器码到剪贴板"""
        clipboard = QApplication.clipboard()
        clipboard.setText(self.machine_code_edit.text())
        QMessageBox.information(self, "成功", "机器码已复制到剪贴板！")
    
    def activate(self):
        """激活软件"""
        activation_code = self.activation_code_edit.text().strip()
        if not activation_code:
            QMessageBox.warning(self, "错误", "请输入激活码！")
            return
        
        if self.license_manager.activate(activation_code):
            QMessageBox.information(self, "成功", "激活成功！感谢您的支持！")
            self.accept()
        else:
            QMessageBox.critical(self, "失败", "激活码无效，请检查后重试！")

class AboutDialog(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("关于")
        self.resize(500, 520)
        self.setStyleSheet("""
            QDialog {
                background-color: #F8FAFC;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # 渐变标题栏
        header = QWidget()
        header.setFixedHeight(100)
        header.setStyleSheet("""
            background: qlineargradient(
                x1:0, y1:0, x2:1, y2:0,
                stop:0 #2563EB,
                stop:1 #7C3AED
            );
        """)
        header_layout = QVBoxLayout(header)
        header_layout.setContentsMargins(30, 20, 30, 20)
        header_layout.setSpacing(5)
        
        title = QLabel(f"{APP_NAME}")
        title.setStyleSheet("""
            font-size: 22px;
            font-weight: 700;
            color: white;
            background: transparent;
        """)
        header_layout.addWidget(title)
        
        subtitle = QLabel(f"{APP_VERSION} · {APP_AUTHOR_CN}")
        subtitle.setStyleSheet("""
            font-size: 13px;
            color: rgba(255, 255, 255, 0.9);
            background: transparent;
        """)
        header_layout.addWidget(subtitle)
        
        layout.addWidget(header)
        
        # 内容区域
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(30, 30, 30, 30)
        content_layout.setSpacing(20)
        
        # 说明文字卡片
        text_card = QFrame()
        text_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        text_shadow = QGraphicsDropShadowEffect()
        text_shadow.setBlurRadius(20)
        text_shadow.setColor(QColor(0, 0, 0, 30))
        text_shadow.setOffset(0, 2)
        text_card.setGraphicsEffect(text_shadow)
        
        text_layout = QVBoxLayout(text_card)
        text_layout.setContentsMargins(20, 20, 20, 20)
        text_layout.setSpacing(10)
        
        txt = QLabel("本软件不收集任何数据和隐私\n如果这个软件对你有帮助，不妨请我喝杯咖啡或奶茶。\n感谢你的认可与支持！")
        txt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        txt.setWordWrap(True)
        txt.setStyleSheet("""
            color: #64748B;
            font-size: 13px;
            line-height: 24px;
        """)
        text_layout.addWidget(txt)
        
        content_layout.addWidget(text_card)
        
        # 二维码卡片
        qr_card = QFrame()
        qr_card.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 12px;
            }
        """)
        
        qr_shadow = QGraphicsDropShadowEffect()
        qr_shadow.setBlurRadius(20)
        qr_shadow.setColor(QColor(0, 0, 0, 30))
        qr_shadow.setOffset(0, 2)
        qr_card.setGraphicsEffect(qr_shadow)
        
        qr_layout = QHBoxLayout(qr_card)
        qr_layout.setContentsMargins(20, 20, 20, 20)
        qr_layout.setSpacing(30)
        qr_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        def make_qr(path, t, icon):
            real_path = resource_path(path)
            w = QWidget(); wl = QVBoxLayout(w); wl.setContentsMargins(0,0,0,0); wl.setSpacing(10); wl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            IMG_SIZE = 140; CONT_SIZE = IMG_SIZE + 10
            l = QLabel(); l.setFixedSize(CONT_SIZE, CONT_SIZE); l.setAlignment(Qt.AlignmentFlag.AlignCenter)
            l.setStyleSheet("background:white; border:1px solid #ddd; border-radius:8px")
            if os.path.exists(real_path): l.setPixmap(QPixmap(real_path).scaled(IMG_SIZE, IMG_SIZE, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            else: l.setText(t); l.setStyleSheet(f"background:#f0f0f0; border:1px solid #ccc; border-radius:8px; color:#999; font-size:14px; qproperty-alignment: AlignCenter;")
            tl = QLabel(t); tl.setAlignment(Qt.AlignmentFlag.AlignCenter); tl.setStyleSheet("color:#333; font-size:14px; font-weight:bold;")
            wl.addWidget(l); wl.addWidget(tl); return w
        qr_layout.addWidget(make_qr("qr1.jpg", "打赏", "💰")); qr_layout.addWidget(make_qr("qr2.jpg", "加好友", "👋"))
        
        content_layout.addWidget(qr_card)
        layout.addWidget(content)

class HandScrollArea(QScrollArea):
    def __init__(self, parent_widget):
        super().__init__()
        self.parent_widget = parent_widget 
        self.setWidgetResizable(True)
        self.setStyleSheet("background-color: #525659; border: none;")
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.is_panning = False; self.last_mouse_pos = QPointF()
        self.enable_interaction = True
    def set_interactive(self, enable):
        self.enable_interaction = enable
        self.setCursor(Qt.CursorShape.OpenHandCursor if enable else Qt.CursorShape.ArrowCursor)
    def wheelEvent(self, event):
        if self.enable_interaction and event.modifiers() == Qt.KeyboardModifier.ControlModifier and self.parent_widget:
            delta = event.angleDelta().y(); 
            if delta > 0: self.parent_widget.zoom_in()
            else: self.parent_widget.zoom_out()
            event.accept()
        else: super().wheelEvent(event)
    def mousePressEvent(self, event):
        if self.enable_interaction and event.button() == Qt.MouseButton.LeftButton:
            self.is_panning = True; self.last_mouse_pos = event.position()
            self.setCursor(Qt.CursorShape.ClosedHandCursor); event.accept()
        else: super().mousePressEvent(event)
    def mouseReleaseEvent(self, event):
        if self.enable_interaction and event.button() == Qt.MouseButton.LeftButton:
            self.is_panning = False; self.setCursor(Qt.CursorShape.OpenHandCursor); event.accept()
        else: super().mouseReleaseEvent(event)
    def mouseMoveEvent(self, event):
        if self.enable_interaction and self.is_panning:
            delta = event.position() - self.last_mouse_pos
            self.last_mouse_pos = event.position()
            self.horizontalScrollBar().setValue(self.horizontalScrollBar().value() - int(delta.x()))
            self.verticalScrollBar().setValue(self.verticalScrollBar().value() - int(delta.y()))
            event.accept()
        else: super().mouseMoveEvent(event)

class AdvancedPreviewArea(QWidget):
    def __init__(self):
        super().__init__()
        self.layout = QVBoxLayout(self); self.layout.setContentsMargins(0,0,0,0); self.layout.setSpacing(0)
        self.scroll_area = HandScrollArea(self)
        self.scroll_area.set_interactive(False)
        self.container = QWidget(); self.container.setStyleSheet("background-color: transparent;")
        self.scroll_layout = QVBoxLayout(self.container); self.scroll_layout.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignTop)
        self.scroll_layout.setSpacing(30); self.scroll_layout.setContentsMargins(20, 20, 20, 20)
        self.scroll_area.setWidget(self.container)
        self.placeholder = QLabel("💡 暂无内容 - 请在左侧添加发票")
        self.placeholder.setStyleSheet("color: #aaa; font-size: 16px; font-weight: bold;")
        self.placeholder.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.scroll_layout.addWidget(self.placeholder)
        self.control_bar = QFrame(); self.control_bar.setObjectName("PreviewControlBar"); self.control_bar.setFixedHeight(45)
        cb_layout = QHBoxLayout(self.control_bar); cb_layout.setContentsMargins(15, 5, 15, 5)
        self.btn_prev = QPushButton(); self.btn_prev.setIcon(Icons.get("prev")); self.btn_prev.setObjectName("IconBtn"); self.btn_prev.setToolTip("上一页"); self.btn_prev.clicked.connect(self.scroll_prev)
        self.lbl_page = QLabel("0 张"); self.lbl_page.setObjectName("PageLabel")
        self.btn_next = QPushButton(); self.btn_next.setIcon(Icons.get("next")); self.btn_next.setObjectName("IconBtn"); self.btn_next.setToolTip("下一页"); self.btn_next.clicked.connect(self.scroll_next)
        cb_layout.addWidget(self.btn_prev); cb_layout.addWidget(self.lbl_page); cb_layout.addWidget(self.btn_next)
        cb_layout.addStretch()
        self.layout.addWidget(self.scroll_area, 1); self.layout.addWidget(self.control_bar)
        self.raw_page_images = []; self.scale_factor = 1.0; self.page_widgets = []
        self.control_bar.setVisible(True)
        self.hq_timer = QTimer(); self.hq_timer.setSingleShot(True); self.hq_timer.timeout.connect(lambda: self.render_pages(True))

    def show_pages(self, page_images):
        self.raw_page_images = page_images; self.render_pages()
        self.lbl_page.setText(f"共 {len(page_images)} 页")

    def render_pages(self, high_quality=False):
        while self.scroll_layout.count():
            item = self.scroll_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        if not self.raw_page_images: self.scroll_layout.addWidget(self.placeholder); return
        for i, img in enumerate(self.raw_page_images):
            page_lbl = QLabel(); pix = QPixmap.fromImage(img)
            view_width = self.scroll_area.viewport().width() - 60
            if view_width < 300: view_width = 300
            mode = Qt.TransformationMode.SmoothTransformation if high_quality else Qt.TransformationMode.FastTransformation
            scaled_pix = pix.scaledToWidth(view_width, mode)
            page_lbl.setPixmap(scaled_pix)
            shadow = QGraphicsDropShadowEffect(); shadow.setBlurRadius(25); shadow.setColor(QColor(0,0,0,120)); shadow.setOffset(0, 8)
            page_lbl.setGraphicsEffect(shadow)
            page_container = QWidget(); pc_layout = QVBoxLayout(page_container); pc_layout.setContentsMargins(0,0,0,0); pc_layout.setSpacing(5)
            pc_layout.addWidget(page_lbl, 0, Qt.AlignmentFlag.AlignCenter)
            num_lbl = QLabel(f"- {i+1} -"); num_lbl.setStyleSheet("color:#888; font-size:11px;"); num_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            pc_layout.addWidget(num_lbl)
            self.scroll_layout.addWidget(page_container)

    def trigger_update(self): self.render_pages(False); self.hq_timer.start(500)
    def scroll_prev(self): self.scroll_area.verticalScrollBar().setValue(max(0, self.scroll_area.verticalScrollBar().value() - 600)); self.trigger_update()
    def scroll_next(self): sb=self.scroll_area.verticalScrollBar(); sb.setValue(min(sb.maximum(), sb.value() + 600)); self.trigger_update()
    def zoom_in(self): pass
    def zoom_out(self): pass
    def resizeEvent(self, e): 
        if self.raw_page_images: self.render_pages()
        super().resizeEvent(e)

class SingleDocViewer(HandScrollArea):
    def __init__(self):
        super().__init__(None) 
        self.set_interactive(True) 
        self.label = QLabel(""); self.label.setAlignment(Qt.AlignmentFlag.AlignCenter); self.setWidget(self.label)
        self.current_pixmap = None; self.zoom_level = 1.0
        self.hq_timer = QTimer(); self.hq_timer.setSingleShot(True); self.hq_timer.timeout.connect(lambda: self.refresh_view(True))

    def wheelEvent(self, event):
        if event.modifiers() == Qt.KeyboardModifier.ControlModifier:
            if event.angleDelta().y() > 0: self.zoom_in()
            else: self.zoom_out()
            event.accept()
        else: super().wheelEvent(event)
    def zoom_in(self): self.zoom_level *= 1.1; self.refresh_view(False); self.hq_timer.start(500)
    def zoom_out(self): self.zoom_level /= 1.1; self.refresh_view(False); self.hq_timer.start(500)
    def set_image(self, img_or_path):
        if isinstance(img_or_path, QImage): self.current_pixmap = QPixmap.fromImage(img_or_path)
        elif isinstance(img_or_path, str) and os.path.exists(img_or_path): self.current_pixmap = QPixmap(img_or_path)
        self.fit_to_window()
    def show_result(self, img): self.set_image(img)
    def fit_to_window(self):
        if not self.current_pixmap: return
        view_size = self.viewport().size()
        w_ratio = (view_size.width() - 40) / self.current_pixmap.width()
        h_ratio = (view_size.height() - 40) / self.current_pixmap.height()
        self.zoom_level = min(w_ratio, h_ratio)
        self.refresh_view(True)
    def refresh_view(self, high_quality=False):
        if not self.current_pixmap: return
        target_w = int(self.current_pixmap.width() * self.zoom_level)
        target_h = int(self.current_pixmap.height() * self.zoom_level)
        mode = Qt.TransformationMode.SmoothTransformation if high_quality else Qt.TransformationMode.FastTransformation
        scaled_pix = self.current_pixmap.scaled(target_w, target_h, Qt.AspectRatioMode.KeepAspectRatio, mode)
        self.label.setPixmap(scaled_pix)
        shadow = QGraphicsDropShadowEffect(); shadow.setBlurRadius(30); shadow.setColor(QColor(0,0,0,180)); shadow.setOffset(0, 10)
        self.label.setGraphicsEffect(shadow)

# ==========================================
# 6. 动态启动页 (Scheme B: 仿毛玻璃)
# ==========================================
class DynamicSplashScreen(QWidget):
    finished = pyqtSignal()
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.WindowStaysOnTopHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setFixedSize(500, 320)
        
        # 主布局
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # 渐变背景卡片
        self.card = QFrame()
        self.card.setStyleSheet("""
            QFrame {
                background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:1,
                    stop:0 #2563EB,
                    stop:1 #7C3AED
                );
                border-radius: 16px;
            }
        """)
        
        # 阴影效果
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(30)
        shadow.setColor(QColor(0, 0, 0, 100))
        shadow.setOffset(0, 8)
        self.card.setGraphicsEffect(shadow)
        main_layout.addWidget(self.card)
        
        # 卡片内容布局
        card_layout = QVBoxLayout(self.card)
        card_layout.setContentsMargins(40, 50, 40, 50)
        card_layout.setSpacing(20)
        
        # 应用名称
        title_lbl = QLabel(APP_NAME)
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_lbl.setStyleSheet("""
            color: white;
            font-weight: 700;
            font-size: 28px;
            letter-spacing: 2px;
            background: transparent;
            border: none;
        """)
        card_layout.addWidget(title_lbl)
        
        # 版本号
        ver_lbl = QLabel(APP_VERSION)
        ver_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ver_lbl.setStyleSheet("""
            color: rgba(255, 255, 255, 0.8);
            font-size: 14px;
            background: transparent;
            border: none;
        """)
        card_layout.addWidget(ver_lbl)
        
        card_layout.addSpacing(10)
        
        # 进度条
        self.progress = QProgressBar()
        self.progress.setFixedHeight(6)
        self.progress.setTextVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: none;
                background-color: rgba(255, 255, 255, 0.2);
                border-radius: 3px;
            }
            QProgressBar::chunk {
                background: qlineargradient(
                    x1:0, y1:0, x2:1, y2:0,
                    stop:0 rgba(255, 255, 255, 0.8),
                    stop:1 rgba(255, 255, 255, 1)
                );
                border-radius: 3px;
            }
        """)
        card_layout.addWidget(self.progress)
        
        # 状态文字
        self.status_lbl = QLabel("正在初始化...")
        self.status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.status_lbl.setStyleSheet("""
            color: rgba(255, 255, 255, 0.9);
            font-size: 13px;
            background: transparent;
            border: none;
        """)
        card_layout.addWidget(self.status_lbl)
        
        card_layout.addStretch()
        
        # 版权信息
        copyright_lbl = QLabel(f"© {APP_AUTHOR_CN}")
        copyright_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        copyright_lbl.setStyleSheet("""
            color: rgba(255, 255, 255, 0.6);
            font-size: 11px;
            background: transparent;
            border: none;
        """)
        card_layout.addWidget(copyright_lbl)
        
        # 扫光效果
        self.shine = QLabel(self.card)
        self.shine.setFixedSize(100, 320)
        self.shine.setStyleSheet("""
            background: qlineargradient(
                x1:0, y1:0, x2:1, y2:0,
                stop:0 rgba(255, 255, 255, 0),
                stop:0.5 rgba(255, 255, 255, 0.3),
                stop:1 rgba(255, 255, 255, 0)
            );
        """)
        self.shine.move(-100, 0)
        
        # 扫光动画
        self.shine_anim = QPropertyAnimation(self.shine, b"pos")
        self.shine_anim.setDuration(2000)
        self.shine_anim.setStartValue(QPoint(-100, 0))
        self.shine_anim.setEndValue(QPoint(500, 0))
        self.shine_anim.setEasingCurve(QEasingCurve.Type.InOutQuad)
        self.shine_anim.setLoopCount(-1)
        self.shine_anim.start()
        
        # 进度定时器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_progress)
        self.timer.start(25)
        self.counter = 0
    def update_progress(self):
        self.counter += 1; self.progress.setValue(self.counter)
        if self.counter < 40: self.status_lbl.setText("正在加载核心组件...")
        elif self.counter < 80: self.status_lbl.setText("正在初始化打印引擎...")
        else: self.status_lbl.setText("准备就绪！")
        if self.counter >= 100: self.timer.stop(); self.finished.emit(); self.close()

# ==========================================
# 7. MainWindow (最后定义)
# ==========================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME} {APP_VERSION}")
        self.resize(1350, 850); self.data = []; self.theme_c = "#555"
        self.temp_files = [] 
        self.preview_timer = QTimer(); self.preview_timer.setSingleShot(True); self.preview_timer.timeout.connect(self.generate_realtime_preview)
        self.current_printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        self.right_panel = None; self.settings_card = None
        self.license_manager = LicenseManager()
        self.init_ui()
        ThemeManager.apply(QApplication.instance())
        self.change_theme("Light")

    def closeEvent(self, event):
        for f in self.temp_files:
            try: os.remove(f)
            except: pass
        super().closeEvent(event)

    def init_ui(self):
        main = QWidget(); self.setCentralWidget(main)
        layout = QHBoxLayout(main); layout.setContentsMargins(15,15,15,15); layout.setSpacing(15)

        # LEFT
        left = QWidget(); left.setFixedWidth(280); lv = QVBoxLayout(left); lv.setContentsMargins(0,0,0,0); lv.setSpacing(12)
        
        # 拖放区域容器
        drop_container = QVBoxLayout(); drop_container.setSpacing(8)
        self.drag = DragArea(); self.drag.dropped.connect(self.add_files)
        drop_hint = QLabel("拖放文件或点击上传"); drop_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        drop_hint.setStyleSheet("color: #94A3B8; font-size: 12px; font-weight: 500;")
        drop_container.addWidget(self.drag); drop_container.addWidget(drop_hint)
        lv.addLayout(drop_container)
        
        # 发票清单标题
        list_title = QLabel("📋 发票清单 (双击修正金额)")
        list_title.setStyleSheet("color: #64748B; font-size: 12px; font-weight: 600; margin-top: 8px;")
        lv.addWidget(list_title)
        
        self.list = QListWidget(); self.list.setIconSize(QSize(40,50)); self.list.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu); self.list.customContextMenuRequested.connect(self.ctx_menu)
        self.list.itemDoubleClicked.connect(self.edit_item); self.list.itemClicked.connect(self.show_single_doc)
        
        tb = QHBoxLayout(); tb.setSpacing(10)
        self.btn_set = QPushButton("设置")
        self.btn_set.setMinimumHeight(44)
        # 平台自适应按钮样式
        if UI_CONFIG.get("use_gradients", True):
            btn_bg = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3B82F6, stop:1 #2563EB);"
            btn_hover = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #2563EB, stop:1 #1D4ED8);"
        else:
            btn_bg = "background: #2563EB;"
            btn_hover = "background: #1D4ED8;"
        self.btn_set.setStyleSheet(f"""
            QPushButton {{
                {btn_bg}
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                {btn_hover}
            }}
        """)
        self.btn_set.clicked.connect(lambda: SettingsDlg(self).exec())
        
        self.btn_del = QPushButton("清空")
        self.btn_del.setMinimumHeight(44)
        if UI_CONFIG.get("use_gradients", True):
            del_bg = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #60A5FA, stop:1 #3B82F6);"
            del_hover = "background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #3B82F6, stop:1 #2563EB);"
        else:
            del_bg = "background: #3B82F6;"
            del_hover = "background: #2563EB;"
        self.btn_del.setStyleSheet(f"""
            QPushButton {{
                {del_bg}
                border: none;
                color: white;
                font-weight: 600;
                font-size: 14px;
                border-radius: 8px;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                {del_hover}
            }}
        """)
        self.btn_del.clicked.connect(self.clear)
        
        tb.addWidget(self.btn_set); tb.addStretch(); tb.addWidget(self.btn_del)
        
        lv.addWidget(self.list); lv.addLayout(tb)
        footer_lbl = QLabel(APP_AUTHOR_CN, alignment=Qt.AlignmentFlag.AlignCenter); footer_lbl.setStyleSheet("color:#999; font-size:11px; margin-top: 10px;")
        lv.addWidget(footer_lbl)

        # MIDDLE
        mid = QWidget(); mv = QVBoxLayout(mid); mv.setContentsMargins(0,0,0,0); mv.setSpacing(0)
        self.stack = QStackedWidget()
        self.word_preview = AdvancedPreviewArea() 
        self.single_viewer = SingleDocViewer() 
        self.stack.addWidget(self.word_preview); self.stack.addWidget(self.single_viewer)
        mv.addWidget(self.stack)

        # RIGHT
        self.right_panel = QWidget(); self.right_panel.setFixedWidth(340)
        rv = QVBoxLayout(self.right_panel); rv.setContentsMargins(0,0,0,0)
        self.settings_card = Card() 
        self.settings_layout = QVBoxLayout(self.settings_card)
        self.settings_layout.setSpacing(15); self.settings_layout.setContentsMargins(20,20,20,20)
        
        # 打印设置标题
        print_title = QLabel("🖨️ 打印设置")
        print_title.setStyleSheet("font-size: 16px; font-weight: 700; color: #1E293B; margin-bottom: 5px;")
        self.settings_layout.addWidget(print_title)
        
        r_pr = QHBoxLayout(); self.cb_pr = QComboBox(); self.cb_pr.addItem("🖥️ 默认打印机/PDF")
        if platform.system() in ["Windows", "Linux"]: 
            for p in QPrinterInfo.availablePrinterNames(): self.cb_pr.addItem(f"🖨️ {p}")
        self.cb_pr.currentIndexChanged.connect(self.on_printer_changed)
        
        self.btn_prop = QPushButton("属性"); self.btn_prop.setObjectName("PropBtn"); self.btn_prop.setFixedSize(60, 30)
        self.btn_prop.clicked.connect(self.open_printer_props)
        self.btn_prop.setEnabled(False)
        r_pr.addWidget(self.cb_pr, 1); r_pr.addWidget(self.btn_prop); self.settings_layout.addLayout(r_pr)

        r_cp = QHBoxLayout(); self.sp_cpy = QSpinBox(); self.sp_cpy.setRange(1,99); self.sp_cpy.setSuffix(" 份")
        self.cb_pap = QComboBox(); self.cb_pap.addItems(["A4", "A5", "B5"]); self.cb_pap.currentTextChanged.connect(self.show_layout_preview)
        r_cp.addWidget(QLabel("份数:")); r_cp.addWidget(self.sp_cpy); r_cp.addWidget(QLabel("纸张:")); r_cp.addWidget(self.cb_pap); self.settings_layout.addLayout(r_cp)
        
        self.settings_layout.addWidget(QLabel("排版模式:"))
        rm = QHBoxLayout(); 
        self.b1 = QToolButton(); self.b1.setObjectName("LayoutCard"); self.b1.setFixedSize(85, 85); self.b1.setIconSize(QSize(72,72))
        self.b2 = QToolButton(); self.b2.setObjectName("LayoutCard"); self.b2.setFixedSize(85, 85); self.b2.setIconSize(QSize(72,72))
        self.b4 = QToolButton(); self.b4.setObjectName("LayoutCard"); self.b4.setFixedSize(85, 85); self.b4.setIconSize(QSize(72,72))
        self.b1.setCheckable(True); self.b2.setCheckable(True); self.b4.setCheckable(True)
        grp=QButtonGroup(self); grp.addButton(self.b1); grp.addButton(self.b2); grp.addButton(self.b4); self.b1.setChecked(True)
        grp.buttonClicked.connect(self.show_layout_preview)
        rm.addWidget(self.b1); rm.addWidget(self.b2); rm.addWidget(self.b4); self.settings_layout.addLayout(rm)

        r_dir = QHBoxLayout()
        self.rd_p = QRadioButton("纵向"); self.rd_l = QRadioButton("横向"); self.rd_l.setChecked(True)
        self.rd_p.toggled.connect(self.update_layout_icons); self.rd_l.toggled.connect(self.update_layout_icons)
        r_dir.addWidget(QLabel("方向:")); r_dir.addWidget(self.rd_p); r_dir.addWidget(self.rd_l); r_dir.addStretch()
        self.settings_layout.addLayout(r_dir)
        
        r_opt = QHBoxLayout()
        self.chk_cut = QCheckBox("显示裁剪辅助线"); self.chk_cut.setChecked(True); self.chk_cut.stateChanged.connect(self.show_layout_preview)
        self.chk_rotate = QCheckBox("强力打印纠偏"); self.chk_rotate.setToolTip("强制旋转90度打印，解决部分打印机方向错误问题")
        r_opt.addWidget(self.chk_cut); r_opt.addSpacing(20); r_opt.addWidget(self.chk_rotate); r_opt.addStretch()
        self.settings_layout.addLayout(r_opt)
        
        rv.addWidget(self.settings_card)

        c3 = Card(); l3 = QVBoxLayout(c3); l3.setSpacing(10); l3.setContentsMargins(20,20,20,20)
        self.lbl_inf = QLabel("0 张发票"); self.lbl_tot = QLabel("¥ 0.00", styleSheet="font-size:22px; font-weight:bold; color:#007AFF")
        l3.addWidget(self.lbl_inf); l3.addWidget(self.lbl_tot)
        self.btn_xls = QPushButton(" 导出 Excel"); self.btn_xls.setIcon(Icons.get("excel")); self.btn_xls.clicked.connect(self.xls); l3.addWidget(self.btn_xls); rv.addWidget(c3)

        self.btn_go = QPushButton(" 开始打印"); self.btn_go.setObjectName("PrimaryBtn"); self.btn_go.setIcon(Icons.get("print", "white")); self.btn_go.setMinimumHeight(50)
        self.btn_go.clicked.connect(self.run); rv.addWidget(self.btn_go)
        
        btn_about = QPushButton(" 关于本软件"); btn_about.clicked.connect(lambda: AboutDialog(self).exec())
        rv.addWidget(btn_about); rv.addStretch()

        layout.addWidget(left); layout.addWidget(mid, 1); layout.addWidget(self.right_panel)
        self.drag.upd("#555")

    def change_theme(self, mode):
        self.theme_c = ThemeManager.apply(QApplication.instance(), mode)
        self.drag.upd(self.theme_c)
        self.btn_set.setIcon(Icons.get("settings", self.theme_c))
        self.btn_del.setIcon(Icons.get("trash", "#d73a49")) 
        self.btn_xls.setIcon(Icons.get("excel", self.theme_c))
        self.btn_go.setIcon(Icons.get("print", "white"))
        self.update_layout_icons() 

    def update_layout_icons(self):
        if not hasattr(self, 'rd_l') or not self.rd_l: return
        try: is_l = self.rd_l.isChecked()
        except RuntimeError: return 
        
        icon_1 = "icon_1x1_l.png" if is_l else "icon_1x1_p.png"
        if os.path.exists(resource_path(icon_1)): self.b1.setIcon(QIcon(resource_path(icon_1)))
        else: self.b1.setIcon(Icons.get("layout_1x1_card", self.theme_c))
        icon_2 = "icon_1x2_l.png" if is_l else "icon_1x2_p.png"
        if os.path.exists(resource_path(icon_2)): self.b2.setIcon(QIcon(resource_path(icon_2)))
        else: self.b2.setIcon(Icons.get("layout_1x2_card_h" if is_l else "layout_1x2_card_v", self.theme_c))
        icon_4 = "icon_2x2_l.png" if is_l else "icon_2x2_p.png"
        if os.path.exists(resource_path(icon_4)): self.b4.setIcon(QIcon(resource_path(icon_4)))
        else: self.b4.setIcon(Icons.get("layout_2x2_card", self.theme_c))
        btn_size = QSize(90, 65) if is_l else QSize(65, 90)
        icon_size = QSize(86, 61) if is_l else QSize(61, 86)
        for btn in [self.b1, self.b2, self.b4]:
            btn.setFixedSize(btn_size)
            btn.setIconSize(icon_size)
        self.show_layout_preview()

    def delete_specific_item(self, item):
        row = self.list.row(item); self.list.takeItem(row); self.data.pop(row); self.calc(); self.show_layout_preview()
    
    def show_single_doc(self, item):
        row = self.list.row(item)
        if row < len(self.data):
            f = self.data[row]['p']
            o = "H" if self.rd_l.isChecked() else "V"
            paper = self.cb_pap.currentText().replace("纸张: ", "") if "纸张: " in self.cb_pap.currentText() else self.cb_pap.currentText()
            doc = PDFEngine.merge([f], "1x1", paper, o, self.chk_cut.isChecked(), out_path=None)
            if doc:
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(3.0, 3.0))
                img = QImage.fromData(pix.tobytes("ppm"))
                
                # [V3.4.0] 单张预览也需要反向旋转修复 (与排版预览逻辑保持一致)
                if o == "H":
                    transform = QTransform()
                    transform.rotate(-90) # 修正: 逆时针90度
                    img = img.transformed(transform)
                    
                self.single_viewer.set_image(img) 
                doc.close() 
            self.stack.setCurrentIndex(1)

    def show_layout_preview(self): self.stack.setCurrentIndex(0); self.trigger_refresh()
    def edit_item(self, item):
        row = self.list.row(item)
        old_val = self.data[row].get('a', 0)
        val, ok = QInputDialog.getDouble(self, "修正金额", "请输入正确金额:", old_val, 0.00, 1000000, 2)
        if ok:
            self.data[row]['a'] = val
            self.data[row]['manually_edited'] = True
            widget = self.list.itemWidget(item)
            if widget:
                widget.update_display(self.data[row])
            self.calc()
    
    def on_printer_changed(self, idx):
        if idx == 0:
            self.btn_prop.setEnabled(False)
            self.btn_go.setText(" 预览 / 生成PDF")
        else:
            self.btn_prop.setEnabled(True)
            p_name = self.cb_pr.currentText().replace("🖨️ ", "")
            self.current_printer = QPrinter(QPrinterInfo.printerInfo(p_name), QPrinter.PrinterMode.HighResolution)
            self.btn_go.setText(f" 打印到: {p_name[:8]}...")

    def open_printer_props(self):
        dlg = QPrintDialog(self.current_printer, self)
        if dlg.exec() == QDialog.DialogCode.Accepted: pass 

    def trigger_refresh(self): self.preview_timer.start(200)
    def generate_realtime_preview(self):
        m="1x1"; m="1x2" if self.b2.isChecked() else m; m="2x2" if self.b4.isChecked() else m
        o="H" if self.rd_l.isChecked() else "V"; 
        
        rotate_preview = (o == "H")

        if hasattr(self, 'current_doc') and self.current_doc: self.current_doc.close(); self.current_doc = None
        gc.collect()
        paper = self.cb_pap.currentText().replace("纸张: ", "") if "纸张: " in self.cb_pap.currentText() else self.cb_pap.currentText()
        self.current_doc = PDFEngine.merge([x['p'] for x in self.data], m, paper, o, self.chk_cut.isChecked(), out_path=None)
        page_imgs = []
        if self.current_doc:
            for page in self.current_doc: 
                pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0)); img = QImage.fromData(pix.tobytes("ppm"))
                if rotate_preview:
                    transform = QTransform()
                    transform.rotate(-90) 
                    img = img.transformed(transform)
                page_imgs.append(img)
        self.word_preview.show_pages(page_imgs)

    def add_files(self, fs):
        logger = logging.getLogger(__name__)
        logger.info(f"开始添加 {len(fs)} 个文件")
        s=QSettings("MySoft","InvoiceMaster"); ak,sk=s.value("ak"),s.value("sk")
        for f in fs:
            logger.info(f"添加文件: {os.path.basename(f)}")
            d = {"p":f, "n":os.path.basename(f), "d":"", "a":0.0, "ext": {}}
            item = QListWidgetItem(self.list); item.setSizeHint(QSize(250, 60))
            widget = InvoiceItemWidget(d, item, self.delete_specific_item); self.list.setItemWidget(item, widget)
            QApplication.processEvents()
            if f.endswith(".pdf"): amt, dt = InvoiceHelper.parse_amount_local(f); d["a"] = amt; d["d"] = dt
            if ak:
                r=InvoiceHelper.ocr(f,ak,sk)
                if r: 
                    if "amount" in r: d["a"]=r["amount"]
                    if "date" in r: d["d"]=r["date"]
                    d["ext"] = r
            self.data.append(d); widget.update_display(d)
        logger.info(f"文件添加完成，共 {len(self.data)} 个发票")
        self.calc(); self.show_layout_preview()
    def calc(self): t=sum(x["a"] for x in self.data); self.lbl_inf.setText(f"{len(self.data)} 张发票"); self.lbl_tot.setText(f"¥ {t:,.2f}")
    def clear(self): self.list.clear(); self.data=[]; self.calc(); self.trigger_refresh()
    def ctx_menu(self, p): m=QMenu(); a=QAction("删除",self); a.triggered.connect(self.del_sel); m.addAction(a); m.exec(self.list.mapToGlobal(p))
    def del_sel(self):
        for r in sorted([self.list.row(i) for i in self.list.selectedItems()], reverse=True): self.list.takeItem(r); self.data.pop(r)
        self.calc(); self.trigger_refresh()
    def xls(self):
        logger = logging.getLogger(__name__)
        if not self.data: return
        
        # 检查激活状态
        info = self.license_manager.get_activation_info()
        if not info['is_activated']:
            if info['remaining_trials'] <= 0:
                # 试用次数用完，必须激活
                QMessageBox.warning(self, "需要激活", "试用次数已用完，请激活软件后继续使用。")
                dialog = ActivationDialog(self, self.license_manager)
                if dialog.exec() != QDialog.DialogCode.Accepted:
                    return
                # 激活成功后继续
            else:
                # 还有试用次数，显示提示并询问是否继续
                remaining = info['remaining_trials']
                reply = QMessageBox.question(
                    self, 
                    "试用提示", 
                    f"您还有 {remaining} 次免费导出机会。\n\n是否继续导出？\n（导出成功后将使用1次机会）",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                if reply != QMessageBox.StandardButton.Yes:
                    return  # 用户选择不继续，直接返回，不扣次数
        
        logger.info(f"开始导出 Excel: {len(self.data)} 条数据")
        
        # 读取上次保存的路径
        s = QSettings("MySoft", "InvoiceMaster")
        last_path = s.value("last_excel_path", os.path.expanduser("~/Desktop/invoice_report.xlsx"))
        
        # 使用上次路径作为默认值
        p, _ = QFileDialog.getSaveFileName(self, "保存 Excel 报表", last_path, "Excel (*.xlsx)")
        if not p: return
        
        # 保存路径供下次使用
        s.setValue("last_excel_path", p)
        
        try:
            # 准备新数据 - 18个专业字段
            new_rows = []
            from datetime import datetime
            import_time = datetime.now().strftime("%Y-%m-%d %H:%M")
            
            for idx, x in enumerate(self.data, 1):
                ext = x.get("ext", {})
                # 处理金额字段,确保是数值类型
                try: amount = float(x.get("a", 0) or 0)
                except: amount = 0
                try: amount_without_tax = float(ext.get("amount_without_tax", "") or 0)
                except: amount_without_tax = ""
                try: tax_amt = float(ext.get("tax_amt", "") or 0)
                except: tax_amt = ""
                
                new_rows.append({
                    "序号": idx,
                    "开票日期": x.get("d", ""), 
                    "发票类型": ext.get("invoice_type", ""),
                    "发票代码": ext.get("code", ""), 
                    "发票号码": ext.get("number", ""), 
                    "校验码": ext.get("check_code", "")[-6:] if ext.get("check_code") else "",  # 后6位
                    "购买方名称": ext.get("buyer", ""),
                    "购买方税号": ext.get("buyer_tax_id", ""),
                    "销售方名称": ext.get("seller", ""), 
                    "销售方税号": ext.get("seller_tax_id", ""),
                    "不含税金额": amount_without_tax,
                    "税率": ext.get("tax_rate", ""),
                    "税额": tax_amt,
                    "价税合计": amount,
                    "商品明细": ext.get("item_name", ""),
                    "备注": ext.get("remark", ""),
                    "导入时间": import_time,
                    "文件路径": x.get("p", "")
                })
            
            new_df = pd.DataFrame(new_rows)
            
            # 如果文件已存在，读取并追加
            if os.path.exists(p):
                try:
                    existing_df = pd.read_excel(p)
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                    is_append = True
                except:
                    combined_df = new_df
                    is_append = False
            else:
                combined_df = new_df
                is_append = False
            
            # 保存到 Excel
            combined_df.to_excel(p, index=False, engine='openpyxl')
            
            # 使用 openpyxl 添加专业样式
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                
                wb = load_workbook(p)
                ws = wb.active
                
                # 定义样式
                header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
                thin_border = Border(
                    left=Side(style='thin', color='E2E8F0'),
                    right=Side(style='thin', color='E2E8F0'),
                    top=Side(style='thin', color='E2E8F0'),
                    bottom=Side(style='thin', color='E2E8F0')
                )
                
                # 应用表头样式
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align
                    cell.border = thin_border
                
                # 冻结首行
                ws.freeze_panes = "A2"
                
                # 设置列宽
                column_widths = [6, 12, 14, 14, 12, 10, 25, 22, 25, 22, 12, 8, 12, 14, 30, 20, 18, 40]
                for idx, width in enumerate(column_widths, 1):
                    if idx <= ws.max_column:
                        ws.column_dimensions[get_column_letter(idx)].width = width
                
                # 检测重复的发票号码（现在是第5列）
                invoice_numbers = {}
                duplicate_rows = set()
                INVOICE_NUM_COL = 5  # 发票号码列
                
                for row_idx in range(2, ws.max_row + 1):
                    invoice_num = ws.cell(row=row_idx, column=INVOICE_NUM_COL).value
                    if invoice_num and str(invoice_num).strip():
                        if invoice_num in invoice_numbers:
                            duplicate_rows.add(row_idx)
                            duplicate_rows.add(invoice_numbers[invoice_num])
                        else:
                            invoice_numbers[invoice_num] = row_idx
                
                # 应用数据行样式(斑马纹 + 重复标记)
                for row_idx in range(2, ws.max_row + 1):
                    is_duplicate = row_idx in duplicate_rows
                    is_zebra = row_idx % 2 == 0
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        if is_duplicate:
                            cell.fill = yellow_fill
                        elif is_zebra:
                            cell.fill = zebra_fill
                        # 金额列右对齐
                        if col_idx in [11, 13, 14]:  # 不含税金额、税额、价税合计
                            cell.alignment = Alignment(horizontal="right")
                
                # 添加底部统计行
                stat_row = ws.max_row + 2
                ws.cell(row=stat_row, column=1, value="统计").font = Font(bold=True)
                ws.cell(row=stat_row, column=2, value=f"共 {ws.max_row - 1} 张发票")
                
                # 计算总金额
                total_amount = 0
                for row_idx in range(2, ws.max_row):
                    try:
                        val = ws.cell(row=row_idx, column=14).value  # 价税合计列
                        if val: total_amount += float(val)
                    except: pass
                ws.cell(row=stat_row, column=14, value=f"¥{total_amount:,.2f}").font = Font(bold=True, color="DC2626")
                
                # 添加工作表保护(安全锁定功能)
                # 允许: 选择单元格、复制、排序、筛选、查找
                # 禁止: 编辑内容、删除行列、修改格式、插入行列
                SHEET_PASSWORD = "InvoiceMaster2024"  # 保护密码
                ws.protection.sheet = True
                ws.protection.password = SHEET_PASSWORD
                ws.protection.enable()
                # 允许的操作
                ws.protection.selectLockedCells = True  # 允许选择锁定单元格
                ws.protection.selectUnlockedCells = True  # 允许选择未锁定单元格
                ws.protection.sort = True  # 允许排序
                ws.protection.autoFilter = True  # 允许筛选
                # 禁止的操作(默认都是False,即禁止)
                ws.protection.formatCells = False  # 禁止格式化单元格
                ws.protection.formatColumns = False  # 禁止格式化列
                ws.protection.formatRows = False  # 禁止格式化行
                ws.protection.insertColumns = False  # 禁止插入列
                ws.protection.insertRows = False  # 禁止插入行
                ws.protection.insertHyperlinks = False  # 禁止插入超链接
                ws.protection.deleteColumns = False  # 禁止删除列
                ws.protection.deleteRows = False  # 禁止删除行
                logger.info("Excel 工作表保护已启用")
                
                wb.save(p)
                
                # 提示信息
                if is_append:
                    msg = f"✅ 已追加 {len(new_df)} 条数据到现有文件！\n"
                    logger.info(f"Excel 追加成功: {p}, 新增 {len(new_df)} 条")
                else:
                    msg = f"✅ 已导出 {len(new_df)} 条数据！\n"
                    logger.info(f"Excel 导出成功: {p}, 共 {len(new_df)} 条")
                
                msg += f"💰 价税合计: ¥{total_amount:,.2f}\n"
                
                if duplicate_rows:
                    msg += f"⚠️ 检测到 {len(duplicate_rows)//2} 组重复发票（已用黄色标记）\n"
                    logger.warning(f"检测到 {len(duplicate_rows)} 条重复发票")
                else:
                    msg += "✓ 未检测到重复发票\n"
                
                msg += "🔒 工作表已保护，仅允许查看和复制"
                
                # 导出成功后，扣除试用次数（如果未激活）
                info = self.license_manager.get_activation_info()
                if not info['is_activated']:
                    self.license_manager.increment_trial_count()
                    logger.info("试用次数已扣除")
                
                QMessageBox.information(self, "导出成功", msg)
                
            except Exception as e:
                # 如果颜色标记失败，至少数据已保存
                logger.warning(f"Excel 颜色标记失败: {str(e)}")
                QMessageBox.information(self, "导出成功", f"数据已导出，但颜色标记失败: {str(e)}")
                
        except Exception as e:
            logger.error(f"Excel 导出失败: {str(e)}", exc_info=True)
            QMessageBox.critical(self, "导出失败", f"错误: {str(e)}")
    def run(self):
        if not self.data: return QMessageBox.warning(self,"Tips","请先添加发票")
        self.btn_go.setText("处理中..."); QApplication.processEvents()
        m="1x1"; m="1x2" if self.b2.isChecked() else m; m="2x2" if self.b4.isChecked() else m
        o="H" if self.rd_l.isChecked() else "V"; out = os.path.expanduser("~/Desktop/Print_Job.pdf")
        
        if out not in self.temp_files: self.temp_files.append(out)
        
        paper = self.cb_pap.currentText().replace("纸张: ", "") if "纸张: " in self.cb_pap.currentText() else self.cb_pap.currentText()
        try:
            # 裁剪线开关直接控制是否显示裁剪线(预览和打印都遵循此设置)
            PDFEngine.merge([x["p"] for x in self.data], m, paper, o, self.chk_cut.isChecked(), out)
            if self.cb_pr.currentIndex() == 0:
                if platform.system() == "Windows": 
                    os.startfile(out, "print") 
                elif platform.system() == "Darwin":  # macOS
                    os.system(f"open '{out}'")
                else:  # Linux/UOS
                    os.system(f"xdg-open '{out}'")
                self.btn_go.setText(" 开始打印")
            else:
                p_name = self.cb_pr.currentText().replace("🖨️ ", ""); copies = self.sp_cpy.value(); self.btn_go.setText(f"正在发送至 {p_name}...")
                
                # [V3.2.0] 传递强力纠偏参数
                force_rotate = self.chk_rotate.isChecked()
                success, msg = PrinterEngine.print_pdf(out, self.current_printer, copies, force_rotate)
                
                if success: 
                    QMessageBox.information(self,"完成","已发送"); self.btn_go.setText(" 开始打印")
                else: QMessageBox.critical(self,"错误",msg); self.btn_go.setText(" 开始打印")
        except Exception as e: self.btn_go.setText("重试"); QMessageBox.critical(self,"Error",str(e))

if __name__ == "__main__":
    # 初始化日志系统
    logger = LogManager.setup_logging()
    logger.info("=" * 60)
    logger.info(f"应用启动 - {APP_NAME} {APP_VERSION}")
    logger.info(f"系统: {platform.system()} {platform.release()}")
    logger.info(f"Python: {sys.version.split()[0]}")
    logger.info(f"日志目录: {LogManager.get_log_directory()}")
    logger.info("=" * 60)
    
    # 全局异常处理
    def exception_hook(exc_type, exc_value, exc_traceback):
        logger.critical(
            "未捕获的异常",
            exc_info=(exc_type, exc_value, exc_traceback)
        )
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
    
    sys.excepthook = exception_hook
    
    try:
        app = QApplication(sys.argv)
        splash = DynamicSplashScreen(); splash.show()
        w = MainWindow(); splash.finished.connect(lambda: [splash.close(), w.show()])
        logger.info("主窗口创建成功")
        exit_code = app.exec()
        logger.info(f"应用正常退出，退出码: {exit_code}")
        sys.exit(exit_code)
    except Exception as e:
        logger.critical(f"应用启动失败: {str(e)}", exc_info=True)
        sys.exit(1)
